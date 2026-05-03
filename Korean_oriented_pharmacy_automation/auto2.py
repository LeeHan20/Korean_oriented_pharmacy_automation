"""
자동화 2번 - 운송장 업로드
===========================================
흐름:
  1.  로젠택배 → 예약관리-주문등록/출력(복수건) → 엑셀저장 요청
  2.  Downloads 에서 '주문등록_출력(복수건)_출력완료' 최신 파일 선택
  3.  Excel → CSV 변환 (데이터 손실 없음)
  4.  홈페이지 주문관리-로젠택배운송장업로드 → 파일선택에 CSV 업로드
  5.  확인 버튼 클릭
"""

import os
import csv
import time
import queue
import threading
import tkinter as tk
from tkinter import scrolledtext, messagebox
from datetime import datetime

import openpyxl

import config
import utils


# ═══════════════════════════════════════════════════════════════════════════════
#  자동화 로직 함수들
# ═══════════════════════════════════════════════════════════════════════════════

def step1_request_rosen_excel(driver):
    """
    로젠택배 예약관리-주문등록/출력(복수건) 페이지에서 엑셀저장 클릭.
    """
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC

    # 로젠택배 탭으로 전환하거나 새 탭 열기
    rosen_found = False
    for handle in driver.window_handles:
        driver.switch_to.window(handle)
        if "ilogen" in driver.current_url:
            rosen_found = True
            break
    if not rosen_found:
        driver.execute_script("window.open(arguments[0]);", config.ROSEN_URL)
        driver.switch_to.window(driver.window_handles[-1])

    wait = WebDriverWait(driver, config.PAGE_LOAD_TIMEOUT)

    # 예약관리 메뉴 (사이드바 - 메인 페이지에 있음)
    menu = wait.until(EC.element_to_be_clickable(
        (By.XPATH, "//*[contains(text(),'예약관리')]")
    ))

    parent = menu.find_element(By.XPATH, "ancestor::*[contains(@class,'toggle-menu')]")
    classes = parent.get_attribute("class")

    if "opened" not in classes:
        menu.click()
    time.sleep(0.4)

    # 주문등록/출력(복수건) (사이드바 - 메인 페이지에 있음)
    sub = wait.until(EC.element_to_be_clickable(
        (By.XPATH,
         "//*[contains(text(),'주문등록') and contains(text(),'복수건')] | "
         "//*[contains(text(),'복수건')]")
    ))
    sub.click()
    time.sleep(1)

    # 콘텐츠는 active 탭의 iframe 안에 있음
    wait.until(EC.frame_to_be_available_and_switch_to_it(
        (By.CSS_SELECTOR, "#ib-contents .ib-tab-contents__item.is-active iframe")
    ))

    # 엑셀저장 버튼 클릭 (onclick="excelDownload();" 으로 직접 찾기)
    before_ts = time.time()
    excel_btn = wait.until(EC.element_to_be_clickable(
        (By.XPATH, "//button[contains(@onclick,'excelDownload')]")
    ))
    excel_btn.click()

    driver.switch_to.default_content()
    return before_ts


def step2_get_rosen_excel(before_ts: float) -> str:
    """Downloads 에서 '주문등록_출력(복수건)_출력완료' 최신 파일 반환."""
    return utils.wait_for_new_file(
        config.DOWNLOAD_DIR,
        "주문등록_출력*복수건*_출력완료*.xls*",
        before_ts
    )


def _cell_to_str(cell) -> str:
    """openpyxl 셀 값을 서버(EUC-KR 파싱) 호환 문자열로 변환."""
    from datetime import datetime as _dt, date as _date
    v = cell.value
    if v is None:
        return ""
    if isinstance(v, float):
        return str(int(v)) if v == int(v) else str(v)
    if isinstance(v, (_dt, _date)):
        return v.strftime("%Y-%m-%d")
    return str(v)


def step3_excel_to_csv(excel_path: str) -> str:
    """
    xlsx → CSV 변환.
    사이트 서버가 EUC-KR로 파싱하므로 EUC-KR 인코딩으로 저장.
    숫자 float → int 변환으로 운송장번호 등 손실 방지.
    """
    csv_path = os.path.splitext(excel_path)[0] + ".csv"

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    with open(csv_path, "w", newline="", encoding="euc-kr", errors="replace") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows():
            writer.writerow([_cell_to_str(cell) for cell in row])

    return csv_path


def step4_upload_to_homepage(driver, csv_path: str):
    """
    홈페이지 주문관리 → 로젠택배운송장업로드 → 파일 업로드.
    """
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC

    # 홈페이지 탭으로 전환하거나 새 탭 열기
    home_found = False
    for handle in driver.window_handles:
        driver.switch_to.window(handle)
        if "ongkihanyak" in driver.current_url:
            home_found = True
            break
    if not home_found:
        driver.execute_script("window.open(arguments[0]);", config.HOMEPAGE_URL)
        driver.switch_to.window(driver.window_handles[-1])

    wait = WebDriverWait(driver, config.PAGE_LOAD_TIMEOUT)

    # 주문관리 버튼은 left(사이드 메뉴) 프레임에 있음
    wait.until(EC.frame_to_be_available_and_switch_to_it("left"))
    order_management_btn = wait.until(EC.element_to_be_clickable(
        (By.XPATH, "//a[@href='../../modules/shop/admin/admin_list.html']")
    ))
    order_management_btn.click()

    # 클릭 후 right 프레임이 주문관리 페이지로 바뀌므로
    # default_content로 빠져나온 뒤 right 프레임에 진입
    driver.switch_to.default_content()
    wait.until(EC.frame_to_be_available_and_switch_to_it("right"))

    # # 로젠택배운송장업로드 클릭 (right 프레임 내 서브메뉴)
    # sub = wait.until(EC.element_to_be_clickable(
    #     (By.XPATH,
    #      "//*[contains(text(),'로젠택배운송장업로드') or "
    #      "contains(text(),'운송장업로드') or "
    #      "contains(text(),'운송장 업로드')]")
    # ))
    # sub.click()

    # 파일선택 input (name="_attach_" 로 특정)
    file_input = wait.until(EC.presence_of_element_located(
        (By.XPATH, "//input[@type='file' and @name='_attach_']")
    ))
    
    file_input.send_keys(os.path.abspath(csv_path))
    time.sleep(0.5)

    driver.switch_to.default_content()


def step5_click_confirm(driver):
    """
    저장하기 버튼 클릭. (right 프레임 안에 있음)
    """
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC

    wait = WebDriverWait(driver, config.PAGE_LOAD_TIMEOUT)
    driver.switch_to.default_content()
    wait.until(EC.frame_to_be_available_and_switch_to_it("right"))
    confirm_btn = wait.until(EC.element_to_be_clickable(
        (By.XPATH,
         "//a[contains(normalize-space(.),'저장하기')]")
    ))

    confirm_btn.click()

    for i in range(2):
        # onchange 에서 confirm() 대화상자가 뜨면 확인 클릭
        try:
            from selenium.webdriver.common.alert import Alert
            Alert(driver).accept()
        except Exception:
            pass

    driver.switch_to.default_content()


def step6_update_order_status(driver, excel_path: str):
    """
    주문등록_출력 파일 S열(주문번호) = 옹기한약 사이트 고유번호와 매칭하여
    주문 상태를 '택배 발송'으로 변경.
    same_address_db.json을 조회하여 병합된 고유번호도 함께 변경.

    HTML 구조 (행 1개 예시):
      <tr>
        <td>
          <input type="checkbox" name="fs_nos[]" value="121485">
          <input type="hidden" id="fs_status_121485" ...>
        </td>
        <td class="cont trans">  ← 1번째 cont trans: 고유번호 표시
        <td class="cont trans" style="...color:#ff6633">  ← 2번째 cont trans: 주문 상태
        ...
      </tr>
    """
    import json
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait, Select
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.common.exceptions import (
        NoSuchElementException, StaleElementReferenceException
    )

    # ── S열에서 주문번호 수집 ─────────────────────────────────────────────────
    load_path = excel_path
    if load_path.lower().endswith(".xls"):
        load_path = utils.xls_to_xlsx(load_path)

    wb = openpyxl.load_workbook(load_path, data_only=True)
    ws = wb.active
    order_numbers = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 19:
            continue
        s_val = row[18]   # S열 (0-indexed: 18)
        if s_val is None:
            continue
        if isinstance(s_val, float):
            num_str = str(int(s_val))
        else:
            num_str = str(s_val).strip()
        if num_str:
            order_numbers.append(num_str)
    wb.close()

    if not order_numbers:
        return

    # ── same_address_db 조회 → 병합된 고유번호 추가 ──────────────────────────
    db_path = config.SAME_ADDRESS_DB_PATH
    if os.path.exists(db_path):
        with open(db_path, "r", encoding="utf-8") as f:
            same_addr_db: dict = json.load(f)
    else:
        same_addr_db = {}

    all_order_numbers: list[str] = []
    for ono in order_numbers:
        all_order_numbers.append(ono)
        if ono in same_addr_db:
            all_order_numbers.extend(same_addr_db[ono])

    # ── 홈페이지 탭으로 전환 ─────────────────────────────────────────────────
    for handle in driver.window_handles:
        driver.switch_to.window(handle)
        if "ongkihanyak" in driver.current_url:
            break

    wait = WebDriverWait(driver, config.PAGE_LOAD_TIMEOUT)
    driver.switch_to.default_content()
    wait.until(EC.frame_to_be_available_and_switch_to_it("right"))

    for order_no in all_order_numbers:
        try:
            # 고유번호(value)가 order_no 인 체크박스 찾기
            checkbox = driver.find_element(
                By.CSS_SELECTOR, f"input[name='fs_nos[]'][value='{order_no}']"
            )
            tr = checkbox.find_element(By.XPATH, "./ancestor::tr[1]")

            # tr 안에 인라인으로 존재하는 select[name='fs_status'] 를 직접 찾아 변경
            # onchange="fs_status_change(this,'고유번호')" 가 자동 저장 처리
            status_select = tr.find_element(By.CSS_SELECTOR, "select[name='fs_status']")
            sel = Select(status_select)

            # value="300" 이 택배발송 옵션
            try:
                sel.select_by_value("300")
            except Exception:
                # fallback: 텍스트로 찾기
                for opt in sel.options:
                    if "택배" in opt.text and "발송" in opt.text:
                        sel.select_by_visible_text(opt.text)
                        break

            # onchange 에서 confirm() 대화상자가 뜨면 확인 클릭
            try:
                from selenium.webdriver.common.alert import Alert
                Alert(driver).accept()
            except Exception:
                pass

            time.sleep(0.3)

        except (NoSuchElementException, StaleElementReferenceException):
            # 해당 주문번호가 목록에 없으면 건너뜀
            pass

    driver.switch_to.default_content()


# ═══════════════════════════════════════════════════════════════════════════════
#  tkinter GUI 앱
# ═══════════════════════════════════════════════════════════════════════════════

class Auto2App:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("자동화 2번 - 운송장 업로드")
        self.root.geometry("480x480")
        self.root.resizable(True, True)

        utils.set_root(root)

        self._q: queue.Queue = queue.Queue()
        self._running = False

        self._build_ui()
        self._poll_queue()

    def _build_ui(self):
        top = tk.Frame(self.root, bg="#1A5276", padx=10, pady=8)
        top.pack(fill="x")
        tk.Label(top, text="자동화 2번 - 운송장 업로드",
                 font=("맑은 고딕", 13, "bold"),
                 bg="#1A5276", fg="white").pack()

        self._status_var = tk.StringVar(value="■ 대기 중")
        tk.Label(self.root, textvariable=self._status_var,
                 font=("맑은 고딕", 10), fg="#2471A3",
                 anchor="w", padx=10).pack(fill="x", pady=(6, 0))

        log_frame = tk.Frame(self.root)
        log_frame.pack(fill="both", expand=True, padx=10, pady=4)
        self._log = scrolledtext.ScrolledText(
            log_frame, height=16, font=("Consolas", 9),
            state="disabled", bg="#FAFAFA"
        )
        self._log.pack(fill="both", expand=True)

        btn_frame = tk.Frame(self.root, pady=8)
        btn_frame.pack()
        self._run_btn = tk.Button(
            btn_frame, text="  시 작  ",
            command=self._start,
            font=("맑은 고딕", 11, "bold"),
            bg="#1A5276", fg="white", padx=18, pady=6,
            relief="flat", cursor="hand2"
        )
        self._run_btn.pack(side="left", padx=6)
        tk.Button(
            btn_frame, text="  닫 기  ",
            command=self.root.destroy,
            font=("맑은 고딕", 11),
            padx=18, pady=6,
            relief="flat", cursor="hand2"
        ).pack(side="left", padx=6)

    def _log_write(self, msg: str):
        self._log.config(state="normal")
        ts = datetime.now().strftime("%H:%M:%S")
        self._log.insert("end", f"[{ts}] {msg}\n")
        self._log.see("end")
        self._log.config(state="disabled")

    def _poll_queue(self):
        while not self._q.empty():
            kind, payload = self._q.get_nowait()
            if kind == "log":
                self._log_write(payload)
            elif kind == "status":
                self._status_var.set(payload)
            elif kind == "done":
                self._running = False
                self._run_btn.config(state="normal")
        self.root.after(80, self._poll_queue)

    def _put(self, kind: str, payload: str):
        self._q.put((kind, payload))

    def _log_msg(self, msg: str):
        self._put("log", msg)
        self._put("status", f"▶ {msg}")

    def _start(self):
        if self._running:
            return
        self._running = True
        self._run_btn.config(state="disabled")
        threading.Thread(target=self._run_all, daemon=True).start()

    def _run_all(self):
        import pythoncom
        pythoncom.CoInitialize()
        driver = None
        try:
            # 1. 로젠택배 엑셀저장
            self._log_msg("1단계: 로젠택배 엑셀저장 요청 중...")
            driver = utils.connect_chrome()
            before_ts = step1_request_rosen_excel(driver)
            self._log_msg("  ✓ 엑셀저장 요청 완료")

            # 2. 파일 대기
            self._log_msg("2단계: 다운로드 파일 대기 중...")
            excel_path = step2_get_rosen_excel(before_ts)
            self._log_msg(f"  ✓ 파일: {os.path.basename(excel_path)}")

            # 3. CSV 변환
            self._log_msg("3단계: CSV 변환 중...")
            csv_path = step3_excel_to_csv(excel_path)
            self._log_msg(f"  ✓ 변환: {os.path.basename(csv_path)}")

            # 4. 홈페이지 업로드
            self._log_msg("4단계: 홈페이지에 운송장 업로드 중...")
            step4_upload_to_homepage(driver, csv_path)
            self._log_msg("  ✓ 파일 선택 완료")

            # 5. 확인 버튼
            self._log_msg("5단계: 확인 버튼 클릭 중...")
            step5_click_confirm(driver)
            self._log_msg("  ✓ 확인 클릭 완료")

            # 6. 주문 상태 → 택배 발송
            self._log_msg("6단계: 주문 상태를 택배 발송으로 변경 중...")
            step6_update_order_status(driver, excel_path)
            self._log_msg("  ✓ 주문 상태 변경 완료")

            self._log_msg("━━━ 자동화 2번 완료 ━━━")
            self._put("status", "✅ 완료")

        except Exception as e:
            self._log_msg(f"오류 발생: {e}")
            self._put("status", "❌ 오류 발생")
            self.root.after(0, lambda e=e: messagebox.showerror(
                "오류", str(e), parent=self.root
            ))
        finally:
            pythoncom.CoUninitialize()
            self._put("done", "")


# ═══════════════════════════════════════════════════════════════════════════════
#  진입점
# ═══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    root = tk.Tk()
    app = Auto2App(root)
    root.mainloop()
