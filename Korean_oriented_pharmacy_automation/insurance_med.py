"""
보험한약 자동화 - insurance_med.py
============================================================
흐름:
  1.  OKOSC 대기처방에서 "보험" 키워드가 있는 행 선택
  2.  약재명/용량을 첩약보험약값대장.xlsx 가격표 시트 B/C열 12행부터 기재
      (괄호 안 내용 제거)
  3.  가격표 E~L열에 공란이 있으면 사용자에게 원산지 확인 요청
  4.  환자명 → 가격표 O6, 연락처 → O7, 주소 → O5
  5.  OKOSC 출력-첩약보험(처방전)-인쇄 클릭
  6.  처방전에서 주민번호(O8), 한의원명(M6), 연락처(M7), 한의사명(M8),
      면허번호(M9), 기관번호(M10), 발급연월일(O10), 발급번호(P10) 추출
  7.  기준처방명(string+alpha+int)의 string 부분으로 처방 선택
  8.  질병분류기호 → 가격표 F5에서 선택
  9.  사용자 기타사항 입력 후 확인
  10. 1일복용팩수 → 가격표 T9, 용법 → 가격표 T10
  11. OKOSC 복용법 → 가격표 T11
  12. 한약재명칭및코드 → 요양급여비용명세서_양식 U16~,
      1회투약량 → Y16~, 1일투여횟수 → Z16~
  13. 총투약횟수 → AA16~
  14. 가감 → AF16~ (감 포함 시 Y열 음수)
  15. 제품코드 → 요양급여비용명세서_양식 R16~
  16. 약재단가 → 요양급여비용명세서_양식 W16~
  17. 사람 확인
  18. 요양급여비용명세서_양식 A~AF → 요양급여명세서모음2 가장 오른쪽 빈 열에 붙여넣기 + 인쇄
  19. 조제내역안내서_양식 A~W → 조제내역안내서양식_모음 가장 오른쪽 빈 열에 붙여넣기
  20. 약재비영수증샘플(실비청구용) A~E → 약재비_영수증_모음 가장 오른쪽 빈 행에 붙여넣기

caution.
  모든 수정 과정에서 시트 보호를 해제하고, 편집 후 다시 보호를 설정함.
"""

import os
import re
import time
import queue
import threading
import warnings
import tkinter as tk
from tkinter import scrolledtext, messagebox
from datetime import datetime

# openpyxl이 EMF 이미지를 읽지 못할 때 발생하는 경고 억제 (데이터에 영향 없음)
warnings.filterwarnings("ignore", message=r".*\.emf.*", category=UserWarning)

import openpyxl
from openpyxl.utils import get_column_letter

import config
import utils


# ─── 시트 이름 상수 ───────────────────────────────────────────────────────────

SHEET_PRICE      = "가격표"
SHEET_YOYANG     = "요양급여비용명세서_양식"
SHEET_YOYANG_COL = "요양급여명세서모음2"
SHEET_JOJE       = "조제내역안내서_양식"
SHEET_JOJE_COL   = "조제내역안내서양식_모음"
SHEET_RECEIPT    = "약재비영수증샘플(실비청구용)"
SHEET_RECEIPT_COL = "약재비_영수증_모음"

# 가격표 시트 한약재 열 (1-based)
PRICE_HERB_COL  = 21   # U열 : 한약재 이름
PRICE_CODE_COL  = 23   # W열 : 제품코드
PRICE_VAL_COL   = 19   # S열 : 약재값
PRICE_START_ROW = 16   # U16부터

# 요양급여비용명세서_양식 열 (1-based)
YY_HERB_COL   = 21   # U열 : 한약재명칭및코드
YY_DOSE1_COL  = 25   # Y열 : 1회투약량
YY_FREQ_COL   = 26   # Z열 : 1일투여횟수
YY_TOTAL_COL  = 27   # AA열 : 총투약횟수
YY_ADD_COL    = 32   # AF열 : 가감
YY_CODE_COL   = 18   # R열 : 제품코드
YY_PRICE_COL  = 23   # W열 : 단가
YY_START_ROW  = 16

# 괄호 제거 정규식
_PAREN_RE = re.compile(r'\s*[\(（][^\)）]*[\)）]')


# ═══════════════════════════════════════════════════════════════════════════════
#  보조 함수
# ═══════════════════════════════════════════════════════════════════════════════

def _remove_parens(text: str) -> str:
    """괄호 및 괄호 안의 내용을 제거합니다."""
    return _PAREN_RE.sub("", text).strip()


def _safe_write(ws, cell_ref: str, value):
    """병합셀이면 anchor(좌상단) 셀에 씁니다."""
    from openpyxl.cell import MergedCell as _MC
    cell = ws[cell_ref]
    if isinstance(cell, _MC):
        for mr in ws.merged_cells.ranges:
            if cell_ref in mr:
                ws.cell(row=mr.min_row, column=mr.min_col).value = value
                return
    else:
        cell.value = value


def _write_cells(wb_path: str, sheet_name: str, writes: dict,
                 log_fn=None, row_writes: list = None):
    """
    시트에 셀 값을 쓁니다. 파일이 열려있으면 COM 폴백을 사용합니다.

    writes:      {"O5": value, "O6": value, ...}  (cell_ref -> value)
    row_writes:  [(row, col, value), ...]           (openpyxl 전용, 즉 COM에서는 writes로 변환)
    """
    def _log(m):
        if log_fn:
            log_fn(m)

    abs_path = os.path.abspath(wb_path)

    def _do_openpyxl():
        wb = openpyxl.load_workbook(abs_path)
        ws = wb[sheet_name]
        ws.protection.sheet = False
        for ref, val in (writes or {}).items():
            _safe_write(ws, ref, val)
        for row, col, val in (row_writes or []):
            ws.cell(row=row, column=col).value = val
        ws.protection.sheet = True
        _log("  Excel 저장 중...")
        wb.save(abs_path)

    def _do_com():
        import win32com.client
        import pythoncom
        pythoncom.CoInitialize()
        try:
            try:
                xl = win32com.client.GetActiveObject("Excel.Application")
            except Exception:
                xl = win32com.client.Dispatch("Excel.Application")
                xl.Visible = False
                xl.DisplayAlerts = False

            abs_lower = abs_path.lower()
            wb_com = None
            for item in xl.Workbooks:
                if item.FullName.lower() == abs_lower:
                    wb_com = item
                    break
            if wb_com is None:
                wb_com = xl.Workbooks.Open(abs_path)

            ws_com = wb_com.Worksheets(sheet_name)
            if ws_com.ProtectContents:
                ws_com.Unprotect()

            for ref, val in (writes or {}).items():
                ws_com.Range(ref).Value = val
            for row, col, val in (row_writes or []):
                # 열 번호를 Excel 열 문자로 변환 (A=1, B=2, ...)
                col_letter = chr(ord('A') + col - 1) if col <= 26 else None
                if col_letter:
                    ws_com.Range(f"{col_letter}{row}").Value = val

            ws_com.Protect()
            wb_com.Save()
            _log("  Excel 저장 완료 (COM 열린 파일 직접 저장)")
        finally:
            pythoncom.CoUninitialize()

    _log("  Excel 파일 열기 중... (12MB 파일, 잠시 기다려주세요)")
    try:
        _do_openpyxl()
    except PermissionError:
        _log("  주의: 파일이 열려있어 COM으로 저장 시도 중...")
        _do_com()


def _unprotect_sheet(wb_path: str, sheet_name: str):
    """win32com으로 시트 보호를 해제합니다."""
    import win32com.client
    import pythoncom

    pythoncom.CoInitialize()
    xl = None
    try:
        xl = win32com.client.Dispatch("Excel.Application")
        xl.Visible = False
        xl.DisplayAlerts = False
        wb = xl.Workbooks.Open(os.path.abspath(wb_path))
        ws = wb.Worksheets(sheet_name)
        if ws.ProtectContents:
            ws.Unprotect()
        wb.Save()
    finally:
        if xl:
            try:
                xl.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()


def _protect_sheet(wb_path: str, sheet_name: str):
    """win32com으로 시트 보호를 다시 설정합니다."""
    import win32com.client
    import pythoncom

    pythoncom.CoInitialize()
    xl = None
    try:
        xl = win32com.client.Dispatch("Excel.Application")
        xl.Visible = False
        xl.DisplayAlerts = False
        wb = xl.Workbooks.Open(os.path.abspath(wb_path))
        ws = wb.Worksheets(sheet_name)
        ws.Protect()
        wb.Save()
    finally:
        if xl:
            try:
                xl.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()


def _find_last_nonempty_col(ws, check_row: int = 1) -> int:
    """check_row 행에서 값이 있는 가장 오른쪽 열 번호(1-based)를 반환합니다."""
    last = 0
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=check_row, column=col).value not in (None, ""):
            last = col
    return last


def _build_herb_lookup(price_ws) -> dict:
    """
    가격표 시트에서 {약재명: (제품코드, 약재값)} 딕셔너리를 생성합니다.
    U열(21)=약재명, W열(23)=제품코드, S열(19)=약재값, 16행부터
    """
    lookup = {}
    for row in range(PRICE_START_ROW, price_ws.max_row + 1):
        name = price_ws.cell(row=row, column=PRICE_HERB_COL).value
        if not name:
            continue
        code  = price_ws.cell(row=row, column=PRICE_CODE_COL).value
        value = price_ws.cell(row=row, column=PRICE_VAL_COL).value
        lookup[str(name).strip()] = (code, value)
    return lookup


def _set_date_field(dlg, hint: str, date_str: str):
    """날짜 필드를 안전하게 설정합니다."""
    import pyautogui
    try:
        field = dlg.child_window(title_re=f".*({hint}).*", control_type="Edit")
        field.set_text(date_str)
    except Exception:
        try:
            field = dlg.child_window(title_re=f".*({hint}).*",
                                     control_type="DateTimePicker")
            rect = field.rectangle()
            pyautogui.click((rect.left + rect.right) // 2,
                            (rect.top + rect.bottom) // 2)
            pyautogui.hotkey('ctrl', 'a')
            pyautogui.write(date_str.replace("-", ""), interval=0.05)
        except Exception:
            pass


def _call_worker(command: str, timeout: int = 30,
                 extra_args: list = None) -> dict:
    """
    okosc_worker.py를 32비트 Python(config.PYTHON32_PATH)으로 실행하고
    JSON 결과를 반환합니다.
    extra_args: command 다음에 추가로 전달할 argv 목록 (예: PDF 경로)
    """
    import subprocess
    import json as _json

    worker_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "okosc_worker.py")
    try:
        import os as _os
        env = _os.environ.copy()
        env["PYTHONIOENCODING"] = "utf-8"
        env["PYTHONUTF8"] = "1"
        cmd_list = [config.PYTHON32_PATH, worker_path, command] + (extra_args or [])
        proc = subprocess.run(
            cmd_list,
            capture_output=True,
            timeout=timeout,
            creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
            env=env,
        )
        out = proc.stdout.decode("utf-8", errors="replace").strip()
        err = proc.stderr.decode("utf-8", errors="replace").strip()
        if err:
            import sys as _sys
            try:
                _sys.stderr.buffer.write(err.encode("utf-8", errors="replace") + b"\n")
                _sys.stderr.buffer.flush()
            except Exception:
                print(err, file=_sys.stderr, flush=True)
        if not out:
            return {"status": "error", "message": err or "worker 출력 없음"}
        return _json.loads(out)
    except subprocess.TimeoutExpired:
        return {"status": "error", "message": f"worker timeout ({timeout}s)"}
    except Exception as e:
        return {"status": "error", "message": str(e)}


# ═══════════════════════════════════════════════════════════════════════════════
#  단계별 함수
# ═══════════════════════════════════════════════════════════════════════════════

def step1_select_insurance_row(log_fn=None) -> dict:
    """
    OKOSC 대기처방 목록에서 "보험" 키워드가 있는 첫 번째 행을 선택합니다.
    okosc_worker.py를 32비트 Python으로 실행하여 UIA 접근 제한을 우회합니다.
    반환: {"search_dlg": okosc_win, "row_idx": int, "row_text": str}
    """
    def log(m):
        if log_fn:
            log_fn(m)

    log("  32비트 worker로 보험 행 탐색 중...")
    result = _call_worker("step1", timeout=40)

    if result.get("status") != "ok":
        raise ValueError(f"보험 행 선택 실패: {result.get('message')}")

    log(f"  보험 행 발견: {result.get('row_text', '')}")

    # 64비트 Python에서도 rectangle()은 동작 → 창 참조만 유지
    okosc_win = utils.find_okosc_app()

    return {
        "search_dlg":      okosc_win,
        "row_idx":         result.get("row_idx", 0),
        "row_text":        result.get("row_text", ""),
        "patient_name":    result.get("patient_name", ""),
        "patient_contact": result.get("patient_contact", ""),
    }



def step2_fill_price_sheet(wb_path: str, search_dlg, log_fn=None) -> list:
    """
    OKOSC 보험 행 클릭 후 우측 패널의 약재명/용량을
    가격표 시트 B12~, C12~에 입력합니다. (괄호 내용 제거)
    반환: [(약재명, 용량), ...]

    TODO: OKOSC 우측 약재 패널 컨트롤 이름 확인 필요.
    """
    def log(m):
        if log_fn:
            log_fn(m)

    # ── 32비트 worker로 약재 목록 파싱 ──────────────────────────────────────
    herbs = []
    log("  32비트 worker로 약재 목록 파싱 중...")
    res = _call_worker("get_herbs")
    if res.get("status") == "ok":
        herbs = [(name, dose) for name, dose in res.get("herbs", [])]
    else:
        log(f"  경고: {res.get('message')}")

    if not herbs:
        log("  경고: 파싱된 약재가 없습니다.")
        return herbs

    log(f"  약재 {len(herbs)}개 파싱 완료")

    # ── 가격표 시트 B/C열 12행~에 기입 ──────────────────────────────────
    rw = [(13 + i, 2, name) for i, (name, _) in enumerate(herbs)] + \
         [(13 + i, 3, dose) for i, (_, dose) in enumerate(herbs)]
    _write_cells(wb_path, SHEET_PRICE, {}, log_fn=log_fn, row_writes=rw)
    log(f"  가격표 B/C열 {len(herbs)}행 기입 완료 (13~{12 + len(herbs)}행)")

    return herbs


def step3_check_origin(wb_path: str, herbs: list, log_fn=None) -> bool:
    """
    가격표 시트에서 약재 기입 행의 E~L열에 공란이 있으면
    사용자에게 원산지 확인을 요청합니다.
    반환: 계속 진행 여부
    """
    def log(m):
        if log_fn:
            log_fn(m)

    wb = openpyxl.load_workbook(wb_path, data_only=True)
    ws = wb[SHEET_PRICE]

    missing_rows = []
    for i in range(len(herbs)):
        r = 12 + i
        for col in range(5, 13):   # E=5 ~ L=12
            if ws.cell(row=r, column=col).value in (None, ""):
                missing_rows.append(r)
                break

    if not missing_rows:
        log("  원산지 확인: 공란 없음")
        return True

    rows_str = ", ".join(str(r) for r in missing_rows)
    return utils.human_review_dialog(
        "원산지 확인 필요",
        f"가격표 시트에 원산지 확인이 필요한 약재가 있습니다.\n",
        ok_text="확인 완료 - 계속",
        cancel_text="중단",
    )


def step4_fill_patient_info(wb_path: str, search_dlg, log_fn=None,
                             prefill_name: str = "", prefill_contact: str = ""):
    """
    OKOSC에서 환자명/연락처/주소를 파싱하여
    가격표 시트 O6(환자명), O7(연락처), O5(주소)에 입력합니다.

    prefill_name/prefill_contact: step1에서 main grid DataItems에서 미리 수집한 값.
    주소는 32비트 worker get_patient (Table[4] item[32])에서 가져옵니다.
    """
    def log(m):
        if log_fn:
            log_fn(m)

    patient_name = prefill_name
    contact      = prefill_contact
    address      = ""

    log("  32비트 worker로 환자 정보 파싱 중...")
    res = _call_worker("get_patient")
    if res.get("status") == "ok":
        address = res.get("address", "")
        if not patient_name:
            patient_name = res.get("name", "")
        if not contact:
            contact = res.get("contact", "")
        clinic_name    = res.get("clinic_name", "")
        clinic_contact = res.get("clinic_contact", "")
    else:
        log(f"  경고: {res.get('message')}")
        clinic_name = clinic_contact = ""

    log(f"  환자명: {patient_name}, 연락처: {contact}, "
        f"주소: {address[:20] + '...' if len(address) > 20 else address}")
    if clinic_name:
        log(f"  한의원명: {clinic_name}, 한의원전화: {clinic_contact}")

    writes = {"O6": patient_name, "O7": contact, "O8": address}
    if clinic_name:
        writes["M6"] = clinic_name
    if clinic_contact:
        writes["M7"] = clinic_contact
    _write_cells(wb_path, SHEET_PRICE, writes, log_fn=log_fn)
    log("  환자 정보 입력 완료 (O6/O7/O8)")


def step5_save_prescription_pdf(search_dlg, log_fn=None) -> str:
    """
    출력 → 첩약보험(처방전) 클릭 → 처방전출력 창 스크린샷 저장 → 창 닫기.
    반환: PNG 이미지 경로 (str) 또는 "" (실패 시)
    """
    def log(m):
        if log_fn:
            log_fn(m)

    img_path = os.path.join(os.path.expanduser("~"), "Downloads",
                            f"presc_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png")

    log("  32비트 worker로 처방전출력 창 캡처 중...")
    res = _call_worker("get_presc_screenshot", extra_args=[img_path], timeout=30)
    if res.get("status") != "ok":
        log(f"  경고: 처방전 캡처 실패 - {res.get('message')}")
        return ""

    actual_path = res.get("img_path", img_path)
    log(f"  처방전 스크린샷 완료: {os.path.basename(actual_path)}")
    return actual_path


def parse_prescription_pdf(pdf_path: str) -> dict:
    """
    처방전 PDF 텍스트를 추출하고 필요한 필드를 파싱합니다.

    반환 dict 키:
      주민번호, 한의원명, 한의원연락처, 한의사이름, 면허번호, 기관번호,
      발급연월일, 발급번호,
      기준처방명, 질병분류기호,
      일복용팩수, 용법,
      herbs: [{"이름": str, "1회투약량": str, "가감": str}, ...]
    """
    try:
        import pdfplumber
    except ImportError:
        return {"_error": "pdfplumber 미설치. pip install pdfplumber 실행 후 재시도하세요."}

    if not pdf_path or not os.path.exists(pdf_path):
        return {"_error": f"PDF 파일이 없습니다: {pdf_path}"}

    full_text = ""
    tables_raw = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            t = page.extract_text() or ""
            full_text += t + "\n"
            for tbl in page.extract_tables():
                tables_raw.append(tbl)

    result = {
        "주민번호": "", "한의원명": "", "한의원연락처": "",
        "한의사이름": "", "면허번호": "", "기관번호": "",
        "발급연월일": "", "발급번호": "",
        "기준처방명": "", "질병분류기호": "",
        "일복용팩수": "", "용법": "",
        "herbs": [],
    }

    # ── 정규식 파싱 ──────────────────────────────────────────────────────────
    def _find(pattern, text=full_text, group=1):
        m = re.search(pattern, text)
        return m.group(group).strip() if m else ""

    result["주민번호"]     = _find(r'(\d{6}-[0-9*]{7})')
    result["발급연월일"]   = _find(r'발급\s*연월일[:\s]*(\d{4}-\d{2}-\d{2})')
    if not result["발급연월일"]:
        dates = re.findall(r'\d{4}-\d{2}-\d{2}', full_text)
        result["발급연월일"] = dates[-1] if dates else ""
    result["발급번호"]     = _find(r'발급\s*번호[:\s]*(-\d{4,6}|-?\d{4,6})')
    result["한의원명"]     = _find(r'요양기관명[:\s]*([^\n\t]+)')
    if not result["한의원명"]:
        result["한의원명"] = _find(r'한의원명[:\s]*([^\n\t]+)')
    result["한의원연락처"] = _find(r'전화번호[:\s]*(0\d{1,2}[-\s]?\d{3,4}[-\s]?\d{4})')
    result["기관번호"]     = _find(r'요양기관\s*기호[:\s]*([A-Z0-9\-]+)')
    result["한의사이름"]   = _find(r'면허\s*(?:의료인\s*)?성명[:\s]*([가-힣]{2,5})')
    if not result["한의사이름"]:
        result["한의사이름"] = _find(r'한의사\s*성명[:\s]*([가-힣]{2,5})')
    result["면허번호"]     = _find(r'면허\s*번호[:\s]*(\d+)')
    result["기준처방명"]   = _find(r'기준처방명[:\s]*([^\n\t]+)')
    result["질병분류기호"] = _find(r'상병\s*(?:코드|기호|분류)[:\s]*([A-Z]\d{2,4}(?:\.\d+)?)')
    if not result["질병분류기호"]:
        result["질병분류기호"] = _find(r'질병\s*분류\s*기호[:\s]*([A-Z]\d{2,4}(?:\.\d+)?)')

    # 1일복용팩수: "1일복용팩수 (n 팩)" 또는 "n팩/일"
    m = re.search(r'1일\s*복용\s*팩\s*수[^\d]*(\d+)\s*팩', full_text)
    result["일복용팩수"] = m.group(1) if m else ""

    result["용법"] = _find(r'용\s*법[:\s]*([^\n]+)')

    # ── 한약재 테이블 파싱 ───────────────────────────────────────────────────
    # 테이블에서 한약재명칭및코드 / 1회투약량 / 가감 열 찾기
    herb_list = []
    HERB_HEADERS = {"한약재", "명칭", "코드", "한약"}
    DOSE_HEADERS = {"1회", "투약량", "용량"}
    ADDED_HEADERS = {"가감", "가", "감"}

    for tbl in tables_raw:
        if not tbl:
            continue
        # 헤더 행 찾기
        header_row_idx = None
        herb_col = dose_col = added_col = None
        for ri, row in enumerate(tbl):
            cells = [str(c or "").strip() for c in row]
            matches = {"herb": None, "dose": None, "added": None}
            for ci, c in enumerate(cells):
                if any(h in c for h in HERB_HEADERS):
                    matches["herb"] = ci
                if any(h in c for h in DOSE_HEADERS):
                    matches["dose"] = ci
                if any(h in c for h in ADDED_HEADERS):
                    matches["added"] = ci
            if matches["herb"] is not None:
                header_row_idx = ri
                herb_col, dose_col, added_col = matches["herb"], matches["dose"], matches["added"]
                break

        if header_row_idx is None:
            continue

        for row in tbl[header_row_idx + 1:]:
            cells = [str(c or "").strip() for c in row]
            if not cells:
                continue
            herb_raw = cells[herb_col] if herb_col is not None and herb_col < len(cells) else ""
            # 한약재명칭및코드에서 이름만 추출 (코드 제거: 영숫자 혼합 제거)
            herb_name = re.sub(r'\s*[A-Za-z0-9]+\s*$', '', herb_raw).strip()
            herb_name = re.sub(r'\([^)]*\)', '', herb_name).strip()  # 괄호 제거
            if not herb_name or not re.search(r'[가-힣]', herb_name):
                continue
            dose  = cells[dose_col]  if dose_col  is not None and dose_col  < len(cells) else ""
            added = cells[added_col] if added_col is not None and added_col < len(cells) else ""
            herb_list.append({"이름": herb_name, "1회투약량": dose, "가감": added})

    # 테이블 파싱 실패 시 텍스트 기반 fallback
    if not herb_list:
        for line in full_text.splitlines():
            # "우슬 8" / "우슬(동우당/국산) 8" 형식
            m = re.match(r'^([가-힣]{2,10}(?:\([^)]*\))?)\s+([\d.]+)\s*(가감|가|감)?', line.strip())
            if m:
                name = re.sub(r'\([^)]*\)', '', m.group(1)).strip()
                herb_list.append({"이름": name, "1회투약량": m.group(2), "가감": m.group(3) or ""})

    result["herbs"] = herb_list
    return result


def _ocr_winrt(img_path: str) -> str:
    """
    Windows 내장 OCR(WinRT)로 이미지에서 한국어 텍스트를 추출합니다.
    pip install winsdk 필요.
    """
    import asyncio

    async def _run():
        from winsdk.windows.media.ocr import OcrEngine
        from winsdk.windows.globalization import Language
        from winsdk.windows.graphics.imaging import BitmapDecoder
        from winsdk.windows.storage import StorageFile

        abs_path = os.path.abspath(img_path)
        file_obj = await StorageFile.get_file_from_path_async(abs_path)
        stream   = await file_obj.open_async(0)       # 0 = FileAccessMode.Read
        decoder  = await BitmapDecoder.create_async(stream)
        bitmap   = await decoder.get_software_bitmap_async()

        lang   = Language("ko")
        engine = OcrEngine.try_create_from_language(lang)
        if engine is None:
            engine = OcrEngine.try_create_from_user_profile_languages()
        if engine is None:
            return ""

        result = await engine.recognize_async(bitmap)
        if result is None:
            return ""
        return "\n".join(line.text for line in result.lines)

    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        return loop.run_until_complete(_run())
    except Exception as e:
        return f"[WinRT OCR 오류: {e}]"
    finally:
        loop.close()


def parse_prescription_image(img_path: str) -> dict:
    """
    처방전출력 창 스크린샷에서 Windows OCR로 텍스트를 추출하고 필드를 파싱합니다.
    parse_prescription_pdf()와 동일한 dict 구조를 반환합니다.
    """
    try:
        import winsdk.windows.media.ocr  # noqa: F401
    except ImportError:
        return {"_error": "winsdk 미설치. pip install winsdk 실행 후 재시도하세요."}

    if not img_path or not os.path.exists(img_path):
        return {"_error": f"이미지 파일이 없습니다: {img_path}"}

    full_text = _ocr_winrt(img_path)
    if full_text.startswith("[WinRT OCR 오류"):
        return {"_error": full_text}

    result = {
        "주민번호": "", "한의원명": "", "한의원연락처": "",
        "한의사이름": "", "면허번호": "", "기관번호": "",
        "발급연월일": "", "발급번호": "",
        "기준처방명": "", "질병분류기호": "",
        "일복용팩수": "", "용법": "",
        "herbs": [],
        "_raw_text": full_text,   # 디버그용
    }

    def _find(pattern, text=full_text, group=1):
        m = re.search(pattern, text)
        return m.group(group).strip() if m else ""

    result["주민번호"]     = _find(r'(\d{6}-[0-9*]{7})')
    result["발급연월일"]   = _find(r'발급\s*연월일[:\s]*(\d{4}[-. ]\d{2}[-. ]\d{2})')
    if not result["발급연월일"]:
        dates = re.findall(r'\d{4}[-. ]\d{2}[-. ]\d{2}', full_text)
        result["발급연월일"] = re.sub(r'[ .]', '-', dates[-1]) if dates else ""
    result["발급번호"]     = _find(r'발급\s*번호[:\s]*(-?\d{4,6})')
    result["한의원명"]     = _find(r'(?:요양기관명|한의원명)[:\s]*([^\n\t]+)')
    result["한의원연락처"] = _find(r'전화번호[:\s]*(0\d{1,2}[-\s]?\d{3,4}[-\s]?\d{4})')
    result["기관번호"]     = _find(r'요양기관\s*기호[:\s]*([A-Z0-9\-]+)')
    result["한의사이름"]   = _find(r'(?:면허|의료인)\s*성명[:\s]*([가-힣]{2,5})')
    result["면허번호"]     = _find(r'면허\s*번호[:\s]*(\d+)')
    result["기준처방명"]   = _find(r'기준처방명[:\s]*([^\n\t]+)')
    # 질병분류기호: 명시적 패턴 우선, 없으면 "K30" 같은 상병코드 직접 탐색
    result["질병분류기호"] = _find(r'(?:상병|질병)\s*(?:코드|기호|분류)[:\s]*([A-Z]\d{2,4}(?:\.\d+)?)')
    if not result["질병분류기호"]:
        result["질병분류기호"] = _find(r'\b([A-Z]\d{2,4}(?:\.\d+)?)\b')

    m = re.search(r'1일\s*복용\s*팩\s*수[^\d]*(\d+)\s*팩', full_text)
    result["일복용팩수"] = m.group(1) if m else ""
    result["용법"] = _find(r'용\s*법[:\s]*([^\n]+)')

    # ── 한약재 파싱 ─────────────────────────────────────────────────────────
    # 처방전 형식: "숙지황 (3299H1AHM) 15 2 1" 또는 "숙지황 15 2 1 가감"
    herb_re = re.compile(
        r'^([가-힣]{2,10})\s*(?:\([^)]*\))?\s+'
        r'([\d.]+)\s+\d+\s+\d+\s*(가감|가|감)?'
    )
    herb_list = []
    for line in full_text.splitlines():
        hm = herb_re.match(line.strip())
        if hm:
            herb_list.append({
                "이름":      hm.group(1),
                "1회투약량": hm.group(2),
                "가감":      hm.group(3) or "",
            })

    result["herbs"] = herb_list
    return result


def step6_extract_prescription_info(wb_path: str, pdf_path: str, log_fn=None) -> dict:
    """
    처방전 스크린샷 이미지를 OCR로 파싱하여 가격표 시트에 입력합니다.
      주민번호   → O9  (README 4번: O9)
      한의원명   → M6
      한의원연락처 → M7
      한의사이름 → M8
      면허번호   → M9
      기관번호   → M10
      발급연월일 → O10
      발급번호   → P10

    반환: 파싱된 처방전 정보 dict (후속 단계에서 재사용)
    """
    def log(m):
        if log_fn:
            log_fn(m)

    log("  처방전 이미지 OCR 파싱 중...")
    fields = parse_prescription_image(pdf_path)

    if "_error" in fields:
        log(f"  경고: {fields['_error']}")
        return fields

    log(f"  주민번호={fields['주민번호']!r}  한의원={fields['한의원명']!r}")
    log(f"  발급일={fields['발급연월일']!r}  발급번호={fields['발급번호']!r}")
    log(f"  기준처방명={fields['기준처방명']!r}  질병분류={fields['질병분류기호']!r}")
    log(f"  한약재 {len(fields['herbs'])}개 파싱됨")

    _write_cells(wb_path, SHEET_PRICE, {
        "O9":  fields["주민번호"],
        "M6":  fields["한의원명"],
        "M7":  fields["한의원연락처"],
        "M8":  fields["한의사이름"],
        "M9":  fields["면허번호"],
        "M10": fields["기관번호"],
        "O10": fields["발급연월일"],
        "P10": fields["발급번호"],
    }, log_fn=log_fn)
    log("  처방전 정보 가격표 입력 완료 (O9/M6~M10/O10/P10)")

    return fields


def step7_select_prescription_name(wb_path: str, presc_fields: dict, log_fn=None):
    """
    PDF 파싱된 기준처방명에서 한글 부분을 추출하여 가격표 B5에 입력합니다.
    """
    def log(m):
        if log_fn:
            log_fn(m)

    presc_name_raw = (presc_fields or {}).get("기준처방명", "")
    if not presc_name_raw:
        log("  경고: 기준처방명을 처방전 PDF에서 찾지 못함")
        return

    m = re.match(r'([가-힣]+)', presc_name_raw)
    if not m:
        log(f"  경고: 기준처방명 파싱 실패: {presc_name_raw}")
        return

    presc_str = m.group(1)
    log(f"  기준처방명: '{presc_name_raw}' → 검색 키: '{presc_str}'")
    _write_cells(wb_path, SHEET_PRICE, {"B5": presc_str}, log_fn=log_fn)
    log(f"  B5 기준처방명 입력 완료: {presc_str}")


def step8_select_disease_code(wb_path: str, presc_fields: dict, log_fn=None):
    """
    PDF 파싱된 질병분류기호를 가격표 F5에 입력합니다.
    """
    def log(m):
        if log_fn:
            log_fn(m)

    disease_code = (presc_fields or {}).get("질병분류기호", "")
    if not disease_code:
        log("  경고: 질병분류기호를 처방전 PDF에서 찾지 못함")
        return

    log(f"  질병분류기호: {disease_code}")
    _write_cells(wb_path, SHEET_PRICE, {"F5": disease_code}, log_fn=log_fn)
    log(f"  F5 질병분류기호 입력 완료: {disease_code}")


def step9_user_misc_input(log_fn=None) -> str:
    """
    사용자가 기타 사항을 입력하고 확인합니다.
    반환: 입력된 기타사항 문자열 (없으면 빈 문자열)
    """
    def log(m):
        if log_fn:
            log_fn(m)

    result = utils.human_text_input_dialog(
        "기타 사항 입력",
        "기타 사항을 입력해 주세요.\n(없으면 빈칸으로 확인)"
    )
    log(f"  기타 사항: {result if result else '(없음)'}")
    return result or ""


def step10_fill_dosage_info(wb_path: str, presc_fields: dict, log_fn=None):
    """
    PDF 파싱된 1일복용팩수 → 가격표 T9, 용법 → T10에 입력합니다.
    """
    def log(m):
        if log_fn:
            log_fn(m)

    pack_count = (presc_fields or {}).get("일복용팩수", "")
    usage_text  = (presc_fields or {}).get("용법", "")

    log(f"  1일복용팩수: {pack_count!r},  용법: {repr(usage_text[:30]) if usage_text else ''}")
    _write_cells(wb_path, SHEET_PRICE, {"T9": pack_count, "T10": usage_text}, log_fn=log_fn)
    log("  T9/T10 입력 완료")


def step11_fill_okosc_dosage(wb_path: str, search_dlg, log_fn=None):
    """
    OKOSC 복용법 내용을 가격표 시트 T11에 입력합니다.
    32비트 worker를 통해 UIA 접근합니다.
    """
    def log(m):
        if log_fn:
            log_fn(m)

    dosage_text = ""
    log("  32비트 worker로 복용법 파싱 중...")
    res = _call_worker("get_dosage")
    if res.get("status") == "ok":
        dosage_text = res.get("dosage", "")
    else:
        log(f"  경고: {res.get('message')}")

    log(f"  복용법: {dosage_text[:40] if dosage_text else '(없음)'}") 

    _unprotect_sheet(wb_path, SHEET_PRICE)
    try:
        wb = openpyxl.load_workbook(wb_path)
        ws = wb[SHEET_PRICE]
        ws["T11"] = dosage_text
        wb.save(wb_path)
        log("  T11 입력 완료")
    finally:
        _protect_sheet(wb_path, SHEET_PRICE)


def step12_14_fill_herb_details(wb_path: str, presc_fields: dict, log_fn=None) -> list:
    """
    PDF 파싱된 한약재 목록으로 요양급여비용명세서_양식 시트를 채웁니다.
      한약재명칭및코드 → U16~  (열 21)
      1회투약량        → Y16~  (열 25, 가감에 "감" 있으면 음수)
      가감             → AF16~ (열 32)
    반환: [(약재명, 1회투약량, "", "", 가감), ...]
    """
    def log(m):
        if log_fn:
            log_fn(m)

    herb_details = []
    herbs = (presc_fields or {}).get("herbs", [])

    if not herbs:
        log("  경고: PDF에서 파싱된 한약재가 없습니다.")
        return herb_details

    for h in herbs:
        herb_details.append((
            h.get("이름", ""),
            h.get("1회투약량", ""),
            "",   # 1일투여횟수 - PDF에 없으면 공란
            "",   # 총투약횟수   - PDF에 없으면 공란
            h.get("가감", ""),
        ))

    log(f"  한약재 세부사항 {len(herb_details)}개 파싱 완료 (PDF 기반)")

    _write_cells(wb_path, SHEET_YOYANG, {}, log_fn=log_fn)  # sheet 존재 확인용 no-op

    try:
        wb = openpyxl.load_workbook(wb_path)
        ws = wb[SHEET_YOYANG]
        for i, (name, dose_once, freq_day, total, addition) in enumerate(herb_details):
            r = YY_START_ROW + i
            ws.cell(row=r, column=YY_HERB_COL).value = name      # U열
            # Y열: 1회투약량. "감" 포함 시 음수
            try:
                dose_val = float(dose_once)
                if "감" in addition:
                    dose_val = -abs(dose_val)
                ws.cell(row=r, column=YY_DOSE1_COL).value = dose_val
            except (ValueError, TypeError):
                ws.cell(row=r, column=YY_DOSE1_COL).value = dose_once
            ws.cell(row=r, column=YY_FREQ_COL).value  = freq_day  # Z열
            ws.cell(row=r, column=YY_TOTAL_COL).value = total     # AA열
            ws.cell(row=r, column=YY_ADD_COL).value   = addition  # AF열
        wb.save(wb_path)
        log("  요양급여비용명세서_양식 U/Y/Z/AA/AF열 입력 완료")
    except Exception as e:
        log(f"  경고: 한약재 세부사항 Excel 저장 실패: {e}")

    return herb_details


def step15_16_fill_code_and_price(wb_path: str, herb_details: list, log_fn=None):
    """
    가격표 U열 한약재 이름을 기준으로 요양급여비용명세서_양식 시트에:
      제품코드(W열 조회) → R16~ (열 18)
      약재단가(S열 조회) → W16~ (열 23)
    를 입력합니다.
    """
    def log(m):
        if log_fn:
            log_fn(m)

    wb_price = openpyxl.load_workbook(wb_path, data_only=True)
    price_ws = wb_price[SHEET_PRICE]
    herb_lookup = _build_herb_lookup(price_ws)

    if not herb_lookup:
        log("  경고: 가격표에서 한약재 조회 테이블 구성 실패")
        return

    log(f"  가격표 한약재 {len(herb_lookup)}종 로드")

    _unprotect_sheet(wb_path, SHEET_YOYANG)
    try:
        wb = openpyxl.load_workbook(wb_path)
        ws = wb[SHEET_YOYANG]
        matched = 0
        for i in range(len(herb_details)):
            r = YY_START_ROW + i
            herb_name = ws.cell(row=r, column=YY_HERB_COL).value
            if not herb_name:
                continue
            herb_name = str(herb_name).strip()
            if herb_name in herb_lookup:
                code, price = herb_lookup[herb_name]
                ws.cell(row=r, column=YY_CODE_COL).value  = code   # R열=18
                ws.cell(row=r, column=YY_PRICE_COL).value = price  # W열=23
                matched += 1
            else:
                log(f"  경고: '{herb_name}' 가격표에 없음")
        wb.save(wb_path)
        log(f"  제품코드/단가 입력 완료: {matched}/{len(herb_details)}건")
    finally:
        _protect_sheet(wb_path, SHEET_YOYANG)


def step18_copy_yoyang_and_print(wb_path: str, log_fn=None):
    """
    요양급여비용명세서_양식 시트의 A~AF열(1~32열)을
    요양급여명세서모음2 시트의 가장 오른쪽 빈 열에 붙여넣고 인쇄합니다.
    (테스트 환경에서는 "출력" 신호만 표시)
    """
    def log(m):
        if log_fn:
            log_fn(m)

    _unprotect_sheet(wb_path, SHEET_YOYANG_COL)
    try:
        wb = openpyxl.load_workbook(wb_path)
        src_ws = wb[SHEET_YOYANG]
        dst_ws = wb[SHEET_YOYANG_COL]

        # A~AF = 열 1~32 데이터 열 단위로 수집
        src_data = []
        for col in range(1, 33):
            col_vals = [
                src_ws.cell(row=row, column=col).value
                for row in range(1, src_ws.max_row + 1)
            ]
            src_data.append(col_vals)

        last_col = _find_last_nonempty_col(dst_ws, check_row=1)
        start_col = last_col + 1

        for offset, col_vals in enumerate(src_data):
            dst_col = start_col + offset
            for row, val in enumerate(col_vals, start=1):
                if val is not None:
                    dst_ws.cell(row=row, column=dst_col).value = val

        wb.save(wb_path)
        log(f"  요양급여명세서모음2 {start_col}열부터 붙여넣기 완료")
    finally:
        _protect_sheet(wb_path, SHEET_YOYANG_COL)

    # 인쇄 (실제 환경에서 win32com 인쇄로 교체)
    log("  [출력] 요양급여비용명세서 인쇄 신호")
    # TODO: 실제 인쇄 구현 (아래 코드 참조)
    # import win32com.client, pythoncom
    # pythoncom.CoInitialize()
    # xl = win32com.client.Dispatch("Excel.Application")
    # xl.Visible = False
    # wb = xl.Workbooks.Open(os.path.abspath(wb_path))
    # wb.Worksheets(SHEET_YOYANG).PrintOut()
    # xl.Quit()
    # pythoncom.CoUninitialize()


def step19_copy_joje(wb_path: str, log_fn=None):
    """
    조제내역안내서_양식 시트의 A~W열(1~23열)을
    조제내역안내서양식_모음 시트의 가장 오른쪽 빈 열에 붙여넣습니다.
    """
    def log(m):
        if log_fn:
            log_fn(m)

    _unprotect_sheet(wb_path, SHEET_JOJE_COL)
    try:
        wb = openpyxl.load_workbook(wb_path)
        src_ws = wb[SHEET_JOJE]
        dst_ws = wb[SHEET_JOJE_COL]

        src_data = []
        for col in range(1, 24):   # A~W = 1~23
            col_vals = [
                src_ws.cell(row=row, column=col).value
                for row in range(1, src_ws.max_row + 1)
            ]
            src_data.append(col_vals)

        last_col = _find_last_nonempty_col(dst_ws, check_row=1)
        start_col = last_col + 1

        for offset, col_vals in enumerate(src_data):
            dst_col = start_col + offset
            for row, val in enumerate(col_vals, start=1):
                if val is not None:
                    dst_ws.cell(row=row, column=dst_col).value = val

        wb.save(wb_path)
        log(f"  조제내역안내서양식_모음 {start_col}열부터 붙여넣기 완료")
    finally:
        _protect_sheet(wb_path, SHEET_JOJE_COL)


def step20_copy_receipt(wb_path: str, log_fn=None):
    """
    약재비영수증샘플(실비청구용) 시트의 A~E열(1~5열)을
    약재비_영수증_모음 시트의 가장 오른쪽 빈 행에 붙여넣습니다.
    """
    def log(m):
        if log_fn:
            log_fn(m)

    _unprotect_sheet(wb_path, SHEET_RECEIPT_COL)
    try:
        wb = openpyxl.load_workbook(wb_path)
        src_ws = wb[SHEET_RECEIPT]
        dst_ws = wb[SHEET_RECEIPT_COL]

        # A~E 데이터 (행 단위)
        src_rows = []
        for row in range(1, src_ws.max_row + 1):
            row_vals = [src_ws.cell(row=row, column=col).value for col in range(1, 6)]
            if any(v is not None for v in row_vals):
                src_rows.append(row_vals)

        last_row = utils.get_last_data_row(dst_ws, check_cols=(1, 2, 3, 4, 5))
        start_row = last_row + 1

        for offset, row_vals in enumerate(src_rows):
            r = start_row + offset
            for col, val in enumerate(row_vals, start=1):
                if val is not None:
                    dst_ws.cell(row=r, column=col).value = val

        wb.save(wb_path)
        log(f"  약재비_영수증_모음 {start_row}행부터 붙여넣기 완료")
    finally:
        _protect_sheet(wb_path, SHEET_RECEIPT_COL)


# ═══════════════════════════════════════════════════════════════════════════════
#  tkinter GUI 앱
# ═══════════════════════════════════════════════════════════════════════════════

class InsuranceMedApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("보험한약 자동화")
        self.root.geometry("520x540")
        self.root.resizable(True, True)

        utils.set_root(root)

        self._q: queue.Queue = queue.Queue()
        self._running = False

        self._build_ui()
        self._poll_queue()

    def _build_ui(self):
        top = tk.Frame(self.root, bg="#1A5276", padx=10, pady=8)
        top.pack(fill="x")
        tk.Label(top, text="보험한약 자동화",
                 font=("맑은 고딕", 13, "bold"),
                 bg="#1A5276", fg="white").pack()

        self._status_var = tk.StringVar(value="■ 대기 중")
        tk.Label(self.root, textvariable=self._status_var,
                 font=("맑은 고딕", 10), fg="#1A5276",
                 anchor="w", padx=10).pack(fill="x", pady=(6, 0))

        log_frame = tk.Frame(self.root)
        log_frame.pack(fill="both", expand=True, padx=10, pady=4)
        self._log = scrolledtext.ScrolledText(
            log_frame, height=18, font=("Consolas", 9),
            state="disabled", bg="#FAFAFA"
        )
        self._log.pack(fill="both", expand=True)

        btn_frame = tk.Frame(self.root, pady=8)
        btn_frame.pack()
        self._run_btn = tk.Button(
            btn_frame, text="  시 작  ",
            command=self._start,
            font=("맑은 고딕", 11, "bold"),
            bg="#1A5276", fg="white", padx=18, pady=6,
            relief="flat", cursor="hand2"
        )
        self._run_btn.pack(side="left", padx=6)
        tk.Button(
            btn_frame, text="  닫 기  ",
            command=self.root.destroy,
            font=("맑은 고딕", 11),
            padx=18, pady=6,
            relief="flat", cursor="hand2"
        ).pack(side="left", padx=6)

    def _log_write(self, msg: str):
        self._log.config(state="normal")
        ts = datetime.now().strftime("%H:%M:%S")
        self._log.insert("end", f"[{ts}] {msg}\n")
        self._log.see("end")
        self._log.config(state="disabled")

    def _poll_queue(self):
        while not self._q.empty():
            kind, payload = self._q.get_nowait()
            if kind == "log":
                self._log_write(payload)
            elif kind == "status":
                self._status_var.set(payload)
            elif kind == "done":
                self._running = False
                self._run_btn.config(state="normal")
        self.root.after(80, self._poll_queue)

    def _put(self, kind: str, payload: str):
        self._q.put((kind, payload))

    def _log_msg(self, msg: str):
        self._put("log", msg)
        self._put("status", f"▶ {msg}")

    def _start(self):
        if self._running:
            return
        self._running = True
        self._run_btn.config(state="disabled")
        threading.Thread(target=self._run_all, daemon=True).start()

    def _run_all(self):
        import pythoncom
        pythoncom.CoInitialize()
        wb_path = config.INSURANCE_WORKBOOK_PATH
        try:
            # 1단계: OKOSC 보험 행 선택
            self._log_msg("1단계: OKOSC 보험 행 선택 중...")
            result = step1_select_insurance_row(log_fn=self._log_msg)
            search_dlg      = result["search_dlg"]
            prefill_name    = result.get("patient_name", "")
            prefill_contact = result.get("patient_contact", "")
            self._log_msg(f"  ✓ 보험 행 선택 완료 "
                          f"(환자명={prefill_name!r}, 전화={prefill_contact!r})")

            # 2단계: 약재명/용량 → 가격표 B/C열 12행~
            self._log_msg("2단계: 가격표 약재명/용량 입력 중...")
            herbs = step2_fill_price_sheet(
                wb_path, search_dlg, log_fn=self._log_msg
            )
            self._log_msg(f"  ✓ {len(herbs)}개 약재 입력 완료")

            # 3단계: 원산지 확인
            self._log_msg("3단계: 원산지 공란 확인 중...")
            ok = step3_check_origin(wb_path, herbs, log_fn=self._log_msg)
            if not ok:
                self._log_msg("━━━ 사용자 중단 ━━━")
                self._put("status", "⏹ 중단됨")
                return

            # 4단계: 환자 정보 → O5/O6/O7
            self._log_msg("4단계: 환자 정보 입력 중...")
            step4_fill_patient_info(wb_path, search_dlg, log_fn=self._log_msg,
                                    prefill_name=prefill_name,
                                    prefill_contact=prefill_contact)
            self._log_msg("  ✓ 환자 정보 입력 완료")

            # 5단계: 처방전출력 창 스크린샷 캡처
            self._log_msg("5단계: 처방전출력 창 캡처 중...")
            presc_path = step5_save_prescription_pdf(search_dlg, log_fn=self._log_msg)

            # 6단계: 처방전 OCR 파싱 → 가격표 입력
            self._log_msg("6단계: 처방전 정보 추출 중...")
            presc_fields = step6_extract_prescription_info(
                wb_path, presc_path, log_fn=self._log_msg
            )
            self._log_msg("  ✓ 처방전 정보 입력 완료")

            # 7단계: 기준처방명 선택
            self._log_msg("7단계: 기준처방명 선택 중...")
            step7_select_prescription_name(
                wb_path, presc_fields, log_fn=self._log_msg
            )

            # 8단계: 질병분류기호 → F5
            self._log_msg("8단계: 질병분류기호 선택 중...")
            step8_select_disease_code(
                wb_path, presc_fields, log_fn=self._log_msg
            )

            # 9단계: 사용자 기타사항 입력
            self._log_msg("9단계: 기타 사항 입력 대기 중...")
            step9_user_misc_input(log_fn=self._log_msg)

            # 10단계: 1일복용팩수 → T9, 용법 → T10
            self._log_msg("10단계: 1일복용팩수/용법 입력 중...")
            step10_fill_dosage_info(
                wb_path, presc_fields, log_fn=self._log_msg
            )

            # 11단계: OKOSC 복용법 → T11
            self._log_msg("11단계: OKOSC 복용법 → T11 입력 중...")
            step11_fill_okosc_dosage(
                wb_path, search_dlg, log_fn=self._log_msg
            )

            # 12~14단계: 한약재 세부사항 → 요양급여비용명세서_양식
            self._log_msg("12~14단계: 한약재 세부사항 입력 중...")
            herb_details = step12_14_fill_herb_details(
                wb_path, presc_fields, log_fn=self._log_msg
            )
            self._log_msg(f"  ✓ {len(herb_details)}개 한약재 입력 완료")

            # 15~16단계: 제품코드/단가 입력
            self._log_msg("15~16단계: 제품코드/단가 입력 중...")
            step15_16_fill_code_and_price(
                wb_path, herb_details, log_fn=self._log_msg
            )

            # 17단계: 사람 확인
            self._log_msg("17단계: 검토 대기 중...")
            ok = utils.human_review_dialog(
                "내용 확인",
                "지금까지 입력된 내용을 확인해 주세요.\n"
                "계속 진행하려면 '계속 진행'을 누르세요.",
                ok_text="계속 진행",
                cancel_text="중단",
            )
            if not ok:
                self._log_msg("━━━ 사용자 중단 ━━━")
                self._put("status", "⏹ 중단됨")
                return

            # 18단계: 요양급여명세서모음2 붙여넣기 + 인쇄
            self._log_msg("18단계: 요양급여명세서모음2 붙여넣기 및 인쇄 중...")
            step18_copy_yoyang_and_print(wb_path, log_fn=self._log_msg)

            # 19단계: 조제내역안내서양식_모음 붙여넣기
            self._log_msg("19단계: 조제내역안내서양식_모음 붙여넣기 중...")
            step19_copy_joje(wb_path, log_fn=self._log_msg)

            # 20단계: 약재비_영수증_모음 붙여넣기
            self._log_msg("20단계: 약재비_영수증_모음 붙여넣기 중...")
            step20_copy_receipt(wb_path, log_fn=self._log_msg)

            self._log_msg("━━━ 보험한약 자동화 완료 ━━━")
            self._put("status", "✅ 완료")

        except Exception as e:
            self._log_msg(f"오류 발생: {e}")
            self._put("status", "❌ 오류 발생")
            self.root.after(0, lambda e=e: messagebox.showerror(
                "오류", str(e), parent=self.root
            ))
        finally:
            pythoncom.CoUninitialize()
            self._put("done", "")


# ═══════════════════════════════════════════════════════════════════════════════
#  진입점
# ═══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    root = tk.Tk()
    app = InsuranceMedApp(root)
    root.mainloop()

