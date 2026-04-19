"""
한약국 자동화 프로그램 - 공통 유틸리티 함수
"""

import os
import re
import glob
import time
import queue
import threading
import tkinter as tk
from tkinter import messagebox
from pathlib import Path
from datetime import datetime, timedelta

import config


# ─── Chrome / Selenium ───────────────────────────────────────────────────────

def connect_chrome():
    """
    원격 디버깅 모드로 실행 중인 Chrome에 연결합니다.
    '크롬_디버깅모드_실행.bat'으로 Chrome을 먼저 실행해야 합니다.
    """
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options

    opts = Options()
    opts.add_experimental_option(
        "debuggerAddress", f"localhost:{config.CHROME_DEBUG_PORT}"
    )
    try:
        driver = webdriver.Chrome(options=opts)
        driver.implicitly_wait(5)
        return driver
    except Exception as e:
        raise ConnectionError(
            f"Chrome 연결 실패.\n"
            f"'크롬_디버깅모드_실행.bat'으로 Chrome을 실행한 뒤 다시 시도하세요.\n"
            f"상세: {e}"
        )


# ─── 파일 다운로드 대기 ───────────────────────────────────────────────────────

def wait_for_new_file(directory: str, pattern: str, before_mtime: float,
                      timeout: int = None) -> str:
    """
    directory 안에서 pattern 에 맞고 before_mtime 이후에 생성된
    최신 파일이 다운로드 완료될 때까지 기다립니다.
    """
    if timeout is None:
        timeout = config.DOWNLOAD_TIMEOUT

    deadline = time.time() + timeout
    while time.time() < deadline:
        candidates = [
            f for f in glob.glob(os.path.join(directory, pattern))
            if os.path.getmtime(f) >= before_mtime
            and not f.endswith(".crdownload")
            and not f.endswith(".tmp")
        ]
        if candidates:
            return max(candidates, key=os.path.getmtime)
        time.sleep(0.5)
    raise TimeoutError(f"파일 다운로드 타임아웃 ({timeout}초). 패턴: {pattern}")


def get_latest_file(directory: str, pattern: str) -> str:
    """directory 에서 pattern 에 맞는 가장 최신 파일 경로 반환."""
    files = glob.glob(os.path.join(directory, pattern))
    if not files:
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {os.path.join(directory, pattern)}")
    return max(files, key=os.path.getmtime)


# ─── Excel 변환 ──────────────────────────────────────────────────────────────

def _parse_html_table_to_xlsx(html_path: str, xlsx_path: str):
    """HTML 테이블(XLS로 위장된 파일)을 파싱해 XLSX로 저장."""
    from html.parser import HTMLParser
    import openpyxl

    class TableParser(HTMLParser):
        def __init__(self):
            super().__init__()
            self.rows = []
            self._cur_row = None
            self._cur_cell = None

        def handle_starttag(self, tag, attrs):
            if tag == 'tr':
                self._cur_row = []
            elif tag in ('td', 'th'):
                self._cur_cell = []

        def handle_endtag(self, tag):
            if tag == 'tr':
                if self._cur_row is not None:
                    self.rows.append(self._cur_row)
                    self._cur_row = None
            elif tag in ('td', 'th'):
                if self._cur_row is not None and self._cur_cell is not None:
                    self._cur_row.append(''.join(self._cur_cell).strip())
                self._cur_cell = None

        def handle_data(self, data):
            if self._cur_cell is not None:
                self._cur_cell.append(data)

    for enc in ('utf-8-sig', 'euc-kr', 'cp949', 'utf-8'):
        try:
            with open(html_path, 'r', encoding=enc, errors='strict') as f:
                content = f.read()
            break
        except (UnicodeDecodeError, LookupError):
            continue
    else:
        with open(html_path, 'r', encoding='utf-8', errors='replace') as f:
            content = f.read()

    parser = TableParser()
    parser.feed(content)

    wb_w = openpyxl.Workbook()
    ws_w = wb_w.active
    for row_data in parser.rows:
        if any(c for c in row_data):
            ws_w.append(row_data)
    wb_w.save(xlsx_path)


def xls_to_xlsx(xls_path: str) -> str:
    """
    XLS 파일을 XLSX 로 변환합니다 (xlrd + openpyxl, COM/Excel 불필요).
    서버가 HTML을 .xls로 내려보내는 경우도 처리합니다.
    원본 XLS 파일은 그대로 유지됩니다.
    """
    import xlrd
    import openpyxl

    xls_path = os.path.abspath(xls_path)
    xlsx_path = os.path.splitext(xls_path)[0] + ".xlsx"

    # 파일 헤더 확인: HTML 위장 XLS 감지 (서버가 HTML을 .xls로 내려보내는 경우)
    with open(xls_path, 'rb') as f:
        header = f.read(512)
    if header.lstrip().startswith(b'<'):
        _parse_html_table_to_xlsx(xls_path, xlsx_path)
        return xlsx_path

    # 진짜 XLS 바이너리
    wb_r = xlrd.open_workbook(xls_path, formatting_info=False)
    wb_w = openpyxl.Workbook()
    wb_w.remove(wb_w.active)

    for sheet_idx in range(wb_r.nsheets):
        ws_r = wb_r.sheet_by_index(sheet_idx)
        ws_w = wb_w.create_sheet(title=ws_r.name)
        for row in range(ws_r.nrows):
            for col in range(ws_r.ncols):
                cell = ws_r.cell(row, col)
                if cell.ctype == xlrd.XL_CELL_DATE:
                    import xlrd.xldate as xldate
                    val = xldate.xldate_as_datetime(cell.value, wb_r.datemode)
                elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                    val = bool(cell.value)
                elif cell.ctype in (xlrd.XL_CELL_ERROR, xlrd.XL_CELL_EMPTY):
                    val = None
                else:
                    val = cell.value
                ws_w.cell(row=row + 1, column=col + 1).value = val

    wb_w.save(xlsx_path)
    return xlsx_path


def open_excel_visible(path: str):
    """
    파일을 보이는 Excel로 열어 GetActiveObject가 잡힐 수 있도록 합니다.
    이미 같은 경로가 열려 있으면 아무 것도 하지 않습니다.
    반환: win32com Workbook 객체
    """
    import win32com.client
    import pythoncom

    pythoncom.CoInitialize()
    abs_path = os.path.abspath(path)

    # 이미 COM으로 접근 가능한 Excel에 열려 있는지 확인
    try:
        xl = win32com.client.GetActiveObject("Excel.Application")
        for wb in xl.Workbooks:
            if os.path.abspath(wb.FullName) == abs_path:
                return wb
        # 같은 앱에 없으면 열기
        xl.Visible = True
        return xl.Workbooks.Open(abs_path)
    except Exception:
        pass

    # Excel 자체가 없으면 새로 띄우기
    xl = win32com.client.Dispatch("Excel.Application")
    xl.Visible = True
    xl.DisplayAlerts = False
    return xl.Workbooks.Open(abs_path)


# ─── OKOSC 통합문서 Excel 찾기 ───────────────────────────────────────────────

def _get_xl_app_from_xlmain(xlmain_hwnd: int):
    """
    XLMAIN 창 핸들에서 Excel Application COM 객체를 가져옵니다.
    GetActiveObject로 접근되지 않는 별도 Excel 프로세스(OKOSC가 연 경우)에서 사용합니다.
    """
    import ctypes
    import win32com.client
    import win32gui
    import pythoncom

    OBJID_NATIVEOM = 0xFFFFFFF0

    # XLMAIN 아래의 EXCEL7 자식 창 찾기
    excel7 = win32gui.FindWindowEx(xlmain_hwnd, 0, "EXCEL7", None)
    if not excel7:
        return None

    # IDispatch IID: {00020400-0000-0000-C000-000000000046} (LE 인코딩)
    IID_BYTES = (ctypes.c_byte * 16)(
        0x00, 0x04, 0x02, 0x00,  # Data1: 0x00020400 LE
        0x00, 0x00,               # Data2: 0x0000 LE
        0x00, 0x00,               # Data3: 0x0000 LE
        0xC0, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x46  # Data4
    )
    obj_ptr = ctypes.c_void_p()
    hr = ctypes.windll.oleacc.AccessibleObjectFromWindow(
        excel7,
        ctypes.c_ulong(OBJID_NATIVEOM),
        ctypes.byref(IID_BYTES),
        ctypes.byref(obj_ptr)
    )
    if hr != 0 or not obj_ptr.value:
        return None

    try:
        p = pythoncom.ObjectFromAddress(obj_ptr.value, pythoncom.IID_IDispatch)
        return win32com.client.Dispatch(p).Application
    except Exception:
        return None


def get_okosc_workbook(wait_new: bool = False, before_names: set = None):
    """
    OKOSC 택배목록 통합문서를 F12 UI 자동화로 임시 파일에 저장합니다.
    OKOSC의 Excel 인스턴스에 COM으로 직접 접근하지 않아 OKOSC 프리징을 방지합니다.
    반환: 저장된 임시 XLSX 파일 경로 (str)
    """
    import win32gui
    import win32con
    import win32clipboard
    import pyautogui

    title_pat = re.compile(r'통합\s*문서\s*\d+')
    temp_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "_okosc_temp.xlsx")
    abs_temp = os.path.abspath(temp_path)
    deadline = time.time() + 20
    f12_attempted = False

    while time.time() < deadline:
        # XLMAIN 창 중 "통합 문서N" 타이틀인 것 찾기
        target_hwnd = None
        def _enum(hwnd, _):
            nonlocal target_hwnd
            if win32gui.GetClassName(hwnd) == "XLMAIN":
                if title_pat.search(win32gui.GetWindowText(hwnd)):
                    target_hwnd = hwnd
        win32gui.EnumWindows(_enum, None)

        if target_hwnd and not f12_attempted:
            f12_attempted = True
            time.sleep(0.8)  # Excel 완전 초기화 대기

            # 기존 임시 파일 삭제
            try:
                if os.path.exists(abs_temp):
                    os.remove(abs_temp)
            except Exception:
                pass

            # 클립보드에 저장 경로 복사 (pyautogui.write 대신 붙여넣기 → IME 우회)
            try:
                win32clipboard.OpenClipboard(0)
                win32clipboard.EmptyClipboard()
                win32clipboard.SetClipboardData(13, abs_temp)  # CF_UNICODETEXT = 13
                win32clipboard.CloseClipboard()
            except Exception:
                pass

            try:
                win32gui.ShowWindow(target_hwnd, win32con.SW_RESTORE)
                win32gui.SetForegroundWindow(target_hwnd)
                time.sleep(0.5)

                pyautogui.hotkey('f12')          # 다른 이름으로 저장
                time.sleep(2.0)                  # 대화상자 열릴 때까지 대기

                # 파일 이름 필드: 전체 선택 후 클립보드 경로 붙여넣기
                pyautogui.hotkey('ctrl', 'a')
                time.sleep(0.15)
                pyautogui.hotkey('ctrl', 'v')
                time.sleep(0.2)
                pyautogui.press('enter')
                time.sleep(1.0)

                # 덮어쓰기·형식 확인 다이얼로그 처리 (Enter로 수락)
                for _ in range(4):
                    time.sleep(0.5)
                    for dlg_name in ("Microsoft Excel", "Excel"):
                        hw = win32gui.FindWindow(None, dlg_name)
                        if hw and win32gui.IsWindowVisible(hw):
                            pyautogui.press('enter')
                            time.sleep(0.4)
            except Exception:
                pass

        # 파일 생성 완료 확인
        if os.path.exists(abs_temp) and os.path.getsize(abs_temp) > 0:
            time.sleep(0.3)   # 쓰기 완료 여유
            return abs_temp

        time.sleep(0.5)

    raise RuntimeError(
        "'통합 문서 N' 형식의 Excel 창을 찾을 수 없습니다.\n"
        "OKOSC에서 택배목록을 먼저 열어주세요."
    )


def list_excel_workbook_names() -> set:
    """현재 열려 있는 모든 Excel 워크북 이름을 반환합니다."""
    import win32com.client
    import pythoncom

    pythoncom.CoInitialize()
    try:
        xl = win32com.client.GetActiveObject("Excel.Application")
        return {wb.Name for wb in xl.Workbooks}
    except Exception:
        return set()


# ─── 셀 색상 판별 ─────────────────────────────────────────────────────────────

def _is_greenish(interior_color_int: int) -> bool:
    """win32com Interior.Color 값이 녹색 계열인지 판별합니다."""
    r = interior_color_int & 0xFF
    g = (interior_color_int >> 8) & 0xFF
    b = (interior_color_int >> 16) & 0xFF
    return g > r and g > b and g > 100


def get_iksan_green_cells(iksan_path: str) -> list:
    """
    익산대장 xlsx에서 녹색 계열 배경의 L열(12번) 셀을 읽어
    [(이름, 전화번호, 주소), ...] 리스트로 반환합니다. (win32com 사용)
    각 셀 형식: "이름 전화번호 주소" (한 셀에 모두 포함)
    """
    import win32com.client
    import pythoncom

    pythoncom.CoInitialize()
    results = []
    xl = None
    try:
        xl = win32com.client.Dispatch("Excel.Application")
        xl.Visible = False
        xl.DisplayAlerts = False
        wb_com = xl.Workbooks.Open(os.path.abspath(iksan_path))
        ws_com = wb_com.Worksheets(1)

        used_rows = ws_com.UsedRange.Rows.Count
        phone_pat = re.compile(r'\d{2,4}[-. ]?\d{3,4}[-. ]?\d{4}')

        for row in range(1, used_rows + 1):
            cell = ws_com.Cells(row, 12)
            val = cell.Value
            if not val:
                continue
            if not _is_greenish(int(cell.Interior.Color)):
                continue

            text = str(val).strip()
            m = phone_pat.search(text)
            if not m:
                continue
            name = text[:m.start()].strip()
            phone = m.group()
            addr = text[m.end():].strip().lstrip(',').strip()
            if name:
                results.append((name, phone, addr))

        wb_com.Close(False)
    finally:
        if xl:
            try:
                xl.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()

    return results


# ─── 익산대장 파일 찾기 ───────────────────────────────────────────────────────

def find_iksan_file(directory: str = None) -> str:
    """'익산대장'으로 시작하는 Excel 파일 경로를 반환합니다."""
    if directory is None:
        directory = config.IKSAN_FILE_DIR

    for ext in ("*.xlsx", "*.xls", "*.xlsm"):
        matches = glob.glob(os.path.join(directory, "익산대장*" + ext.lstrip("*")))
        if matches:
            return matches[0]
    raise FileNotFoundError(
        f"익산대장 파일을 찾을 수 없습니다.\n경로: {directory}"
    )


# ─── Excel 마지막 데이터 행 ───────────────────────────────────────────────────

def get_last_data_row(ws, check_cols=(1,)) -> int:
    """
    지정한 열(check_cols, 1-based) 중 하나라도 값이 있는
    가장 마지막 행 번호를 반환합니다. 데이터가 없으면 0.
    """
    for row in range(ws.max_row, 0, -1):
        if any(ws.cell(row=row, column=c).value not in (None, "") for c in check_cols):
            return row
    return 0


# ─── 날짜 유틸 ───────────────────────────────────────────────────────────────

def get_search_dates():
    """(시작일 문자열, 오늘 문자열) 반환. 형식: YYYY-MM-DD"""
    today = datetime.now()
    start = today - timedelta(days=config.DATE_RANGE_DAYS)
    return start.strftime("%Y-%m-%d"), today.strftime("%Y-%m-%d")


# ─── 인간 검토 다이얼로그 ─────────────────────────────────────────────────────

def human_text_input_dialog(title: str, prompt: str,
                            ok_text="확인", cancel_text="취소") -> str:
    """
    tkinter 텍스트 입력 다이얼로그를 메인 스레드에서 표시합니다.
    입력 문자열 반환, 취소 시 빈 문자열 반환.
    """
    result_holder = [""]
    event = threading.Event()

    def _show():
        win = tk.Toplevel()
        win.title(title)
        win.grab_set()
        win.resizable(False, False)

        tk.Label(win, text=prompt, font=("맑은 고딕", 10),
                 wraplength=400, justify="left", padx=20, pady=(15, 5)).pack()

        entry_var = tk.StringVar()
        entry = tk.Entry(win, textvariable=entry_var,
                         font=("맑은 고딕", 10), width=40)
        entry.pack(padx=20, pady=(0, 10))
        entry.focus_set()

        btn_frame = tk.Frame(win)
        btn_frame.pack(pady=(0, 15))

        def on_ok():
            result_holder[0] = entry_var.get()
            win.destroy()
            event.set()

        def on_cancel():
            result_holder[0] = ""
            win.destroy()
            event.set()

        tk.Button(btn_frame, text=ok_text, command=on_ok,
                  bg="#4CAF50", fg="white",
                  font=("맑은 고딕", 10, "bold"), padx=15, pady=5).pack(side="left", padx=5)
        tk.Button(btn_frame, text=cancel_text, command=on_cancel,
                  font=("맑은 고딕", 10), padx=15, pady=5).pack(side="left", padx=5)

        win.bind("<Return>", lambda _: on_ok())
        win.protocol("WM_DELETE_WINDOW", on_cancel)

    _ROOT.after(0, _show)
    event.wait()
    return result_holder[0]


def human_review_dialog(title: str, message: str,
                        ok_text="계속 진행", cancel_text="중단") -> bool:
    """
    tkinter 다이얼로그를 메인 스레드에서 표시합니다.
    OK → True, Cancel → False
    """
    result_holder = [None]
    event = threading.Event()

    def _show():
        win = tk.Toplevel()
        win.title(title)
        win.grab_set()
        win.resizable(False, False)

        tk.Label(win, text=message, font=("맑은 고딕", 10),
                 wraplength=400, justify="left", padx=20, pady=20).pack()

        btn_frame = tk.Frame(win)
        btn_frame.pack(pady=(0, 15))

        def on_ok():
            result_holder[0] = True
            win.destroy()
            event.set()

        def on_cancel():
            result_holder[0] = False
            win.destroy()
            event.set()

        tk.Button(btn_frame, text=ok_text, command=on_ok,
                  bg="#4CAF50", fg="white",
                  font=("맑은 고딕", 10, "bold"), padx=15, pady=5).pack(side="left", padx=5)
        tk.Button(btn_frame, text=cancel_text, command=on_cancel,
                  font=("맑은 고딕", 10), padx=15, pady=5).pack(side="left", padx=5)

        win.protocol("WM_DELETE_WINDOW", on_cancel)

    # 메인 스레드에서 창 표시 (caller가 _ROOT 를 전달해야 함)
    _ROOT.after(0, _show)
    event.wait()
    return result_holder[0]


# ─── 공유 tkinter 루트 참조 ──────────────────────────────────────────────────
# 각 auto*.py 에서 실행 시 _ROOT 를 실제 root 로 교체해야 합니다.
_ROOT: tk.Tk = None


def set_root(root: tk.Tk):
    global _ROOT
    _ROOT = root


# ─── OKOSC 창 찾기 ───────────────────────────────────────────────────────────

def find_okosc_app():
    """
    pywinauto를 사용해 OKOSC 애플리케이션 창을 찾습니다.
    반환: pywinauto WindowSpecification (win32 backend)
    """
    import win32gui
    import win32con
    from pywinauto import Application

    hwnd_list = []
    def enum_cb(hwnd, _):
        title = win32gui.GetWindowText(hwnd)
        cls = win32gui.GetClassName(hwnd)
        if any(kw in title for kw in config.OKOSC_WINDOW_KEYWORDS) and 'WindowsForms' in cls:
            hwnd_list.append(hwnd)
    win32gui.EnumWindows(enum_cb, None)

    if not hwnd_list:
        raise RuntimeError(
            "OKOSC 프로그램 창을 찾을 수 없습니다.\n"
            "OKOSC를 먼저 실행해 주세요.\n"
            f"찾는 키워드: {config.OKOSC_WINDOW_KEYWORDS}"
        )

    hwnd = hwnd_list[0]
    win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
    win32gui.SetForegroundWindow(hwnd)
    time.sleep(0.3)

    app = Application(backend="win32").connect(handle=hwnd)
    return app.window(handle=hwnd)


def print_okosc_controls():
    """
    OKOSC 창의 컨트롤 식별자를 출력합니다 (개발/디버깅용).
    터미널에서 실행: python -c "from utils import print_okosc_controls; print_okosc_controls()"
    """
    w = find_okosc_app()
    w.print_control_identifiers()
