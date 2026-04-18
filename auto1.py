"""
자동화 1번 - 택배 주문 처리
===========================================
흐름:
  1.  홈페이지에서 택배 엑셀 다운로드 (조제 중, 익산점)
  2.  XLS → XLSX 변환
  3.  1행 삭제
  4.  G열 '즉납' → '4400' 교체
  5.  OKOSC 처방전검색 → 택배목록 Excel 가져오기
  6.  통합문서 E~I → 택배관리 A~E (이어붙이기)
  7.  통합문서 A, C~D → 택배관리 K, L~M
  8.  새 행의 F=1, G=4400, H=한약 채우기
  9.  D/E열 번호 정규화
  10-11. 익산대장 연한녹색 셀 파싱 → 택배관리 마지막 행에 추가
  12. 사람 검토 (OK → 계속)
  13-14. 로젠택배 사이트 → 예약관리-주문등록/출력(복수건)
  15. 택배관리 파일 업로드
  16. 사람 검토 (오류 열 확인)
"""

import os
import re
import time
import queue
import threading
import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox
from datetime import datetime

import openpyxl
from openpyxl.utils import get_column_letter

import config
import utils


# ═══════════════════════════════════════════════════════════════════════════════
#  자동화 로직 함수들
# ═══════════════════════════════════════════════════════════════════════════════

def step1_download_excel(driver) -> str:
    """
    홈페이지에서 택배 엑셀 다운로드.
    반환: 다운로드된 XLS 파일 경로
    """
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.support.ui import Select

    # 다운로드 직전 시각 기록 (새 파일 감지용)
    before_ts = time.time()

    # 이미 열린 홈페이지 탭으로 전환 (없으면 현재 탭에서 이동)
    for handle in driver.window_handles:
        driver.switch_to.window(handle)
        if "ongkihanyak" in driver.current_url:
            break
    else:
        driver.switch_to.window(driver.window_handles[0])
        driver.get(config.HOMEPAGE_URL)
        WebDriverWait(driver, config.PAGE_LOAD_TIMEOUT).until(
            EC.url_contains("ongkihanyak")
        )

    wait = WebDriverWait(driver, config.PAGE_LOAD_TIMEOUT)

    # 주문관리 버튼은 left(사이드 메뉴) 프레임에 있음
    wait.until(EC.frame_to_be_available_and_switch_to_it("left"))
    order_management_btn = wait.until(EC.element_to_be_clickable(
        (By.XPATH, "//a[@href='../../modules/shop/admin/admin_list.html']")
    ))
    order_management_btn.click()

    # 클릭 후 right 프레임이 주문관리 페이지로 바뀌므로
    # default_content로 빠져나온 뒤 right 프레임에 진입
    driver.switch_to.default_content()
    wait.until(EC.frame_to_be_available_and_switch_to_it("right"))

    # 조제중 선택 (SELECT name=fs_status)
    sel_status = wait.until(EC.presence_of_element_located(
        (By.CSS_SELECTOR, "select[name='fs_status']")
    ))
    Select(sel_status).select_by_visible_text(config.ORDER_STATUS)

    # 익산점 선택 (SELECT name=store)
    sel_store = driver.find_element(By.CSS_SELECTOR, "select[name='store']")
    Select(sel_store).select_by_visible_text(config.BRANCH_NAME)

    time.sleep(0.3)

    # 택배 [엑셀다운] 클릭 - onclick에 excel_down 이 있는 링크
    dl_btn = wait.until(EC.element_to_be_clickable(
        (By.XPATH, "//a[contains(@onclick,'excel_down')]")
    ))
    dl_btn.click()

    driver.switch_to.default_content()

    # 다운로드 완료 대기 (택배관리_*.xls)
    xls_path = utils.wait_for_new_file(
        config.DOWNLOAD_DIR,
        "택배관리_*.xls*",
        before_ts
    )
    return xls_path


def step2_convert_xls(xls_path: str) -> str:
    """XLS → XLSX 변환."""
    return utils.xls_to_xlsx(xls_path)


def step3_remove_first_row(xlsx_path: str):
    """XLSX 파일의 1행(헤더 위 빈 행 또는 불필요 행) 삭제."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    ws.delete_rows(1)
    wb.save(xlsx_path)
    wb.close()


def step4_clean_column_g(xlsx_path: str):
    """G열에서 '즉납' 텍스트를 찾아 '4400'으로 교체."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    for row in ws.iter_rows(min_col=7, max_col=7):
        cell = row[0]
        if cell.value and "즉납" in str(cell.value):
            cell.value = config.DELIVERY_G_VALUE
    wb.save(xlsx_path)
    wb.close()


def step5_automate_okosc() -> object:
    """
    OKOSC에서 날짜 범위 설정 → 검색 → 택배목록 클릭.
    반환: '통합 문서 N' win32com 워크북 객체
    """
    import pyautogui
    import psutil
    from pywinauto.keyboard import send_keys

    start_date, end_date = utils.get_search_dates()  # YYYY-MM-DD

    # ── OKOSC 창 찾기 ──────────────────────────────────────────────────────────
    okosc_win = utils.find_okosc_app()
    okosc_win.set_focus()
    time.sleep(0.5)

    # ── 검색 기준 → 진행상태 설정 (auto_id=ulCboSearch) ──────────────────────
    try:
        search_type_ctrl = okosc_win.child_window(auto_id="ulCboSearch")
        rect = search_type_ctrl.rectangle()
        pyautogui.click((rect.left + rect.right) // 2, (rect.top + rect.bottom) // 2)
        time.sleep(0.15)
        send_keys('^a')
        send_keys('진행상태')
        send_keys('{ENTER}')
        time.sleep(0.3)
    except Exception:
        pass

    def _set_ultra_date(auto_id, date_str):
        """
        UltraDateTimeEditor에 날짜 입력.
        컨트롤 클릭 → End → Backspace 8번(20xx-xx-xx → 20) →
        년 2자리 + 월 2자리 + 일 2자리 입력.
        """
        y, m, d = date_str.split("-")
        yy = y[2:]   # 년도 뒤 2자리 (예: "2026" → "26")
        ctrl = okosc_win.child_window(auto_id=auto_id)
        rect = ctrl.rectangle()
        cx = (rect.left + rect.right) // 2
        cy = (rect.top + rect.bottom) // 2

        pyautogui.click(cx, cy)
        time.sleep(0.2)
        send_keys('{END}')           # 커서를 일 파트 끝으로
        time.sleep(0.1)
        for _ in range(8):           # 일(2)+-(1)+월(2)+-(1)+년뒤2자리(2) = 8
            send_keys('{BACKSPACE}')
            time.sleep(0.05)
        time.sleep(0.1)
        # 년 2자리 → 월 2자리 → 일 2자리
        for ch in yy + m + d:
            send_keys(ch)
            time.sleep(0.05)
        send_keys('{TAB}')
        time.sleep(0.3)

    _set_ultra_date("ulDteSearchStart", start_date)
    _set_ultra_date("ulDteSearchEnd", end_date)

    # TAB 후 드롭다운이 열렸을 수 있으므로 ESC로 닫기
    send_keys('{ESC}')
    time.sleep(0.2)
    okosc_win.set_focus()
    time.sleep(0.2)

    # ── 진행상태 → 조제 선택 (auto_id=ulCboSearchCBJState) ───────────────────
    try:
        state_ctrl = okosc_win.child_window(auto_id="ulCboSearchCBJState")
        rect = state_ctrl.rectangle()
        pyautogui.click((rect.left + rect.right) // 2, (rect.top + rect.bottom) // 2)
    except Exception:
        # 못 찾으면 title로 재시도
        try:
            state_ctrl = okosc_win.child_window(title_re="조제|대기|완료|취소",
                                                class_name_re=".*UltraComboEditor.*")
            rect = state_ctrl.rectangle()
            pyautogui.click((rect.left + rect.right) // 2, (rect.top + rect.bottom) // 2)
        except Exception:
            pass  # 이미 조제로 되어 있을 경우 무시
    time.sleep(0.15)
    send_keys('^a')
    send_keys('조제')
    send_keys('{ENTER}')
    time.sleep(0.3)

    # ── 검색 버튼 클릭 (auto_id=ulBtnSearchCBJ) ──────────────────────────────
    okosc_win.child_window(auto_id="ulBtnSearchCBJ").click_input()
    time.sleep(2)

    # ── 택배목록 버튼 클릭 (auto_id=ulBtnTekBe) → Excel 파일 열림 ───────────
    okosc_win.child_window(auto_id="ulBtnTekBe").click_input()

    # ── 통합문서가 열릴 때까지 대기 후 COM 연결 ──────────────────────────────
    okosc_wb = utils.get_okosc_workbook()
    return okosc_wb


def step6_7_8_paste_okosc_data(xlsx_path: str, okosc_wb) -> tuple:
    """
    통합문서 데이터를 택배관리 파일에 붙여넣기.
    - 통합문서 E~I → 택배관리 A~E (이어붙이기)
    - 통합문서 A, C~D → 택배관리 K, L~M
    - 새 행 F=1, G=4400, H=한약
    반환: (새 데이터 시작 행, 새 데이터 끝 행)
    """
    # ── 통합문서에서 데이터 읽기 (win32com) ──────────────────────────────────
    rows_data = []
    okosc_ws = okosc_wb.Worksheets(1)
    r = 2  # 1행은 헤더로 가정
    while True:
        val_a = okosc_ws.Cells(r, 1).Value
        if val_a is None and okosc_ws.Cells(r, 5).Value is None:
            break
        rows_data.append({
            "A": okosc_ws.Cells(r, 1).Value,    # K로 갈 값
            "C": okosc_ws.Cells(r, 3).Value,    # L로 갈 값
            "D": okosc_ws.Cells(r, 4).Value,    # M으로 갈 값
            "E": okosc_ws.Cells(r, 5).Value,    # 택배관리 A
            "F": okosc_ws.Cells(r, 6).Value,    # 택배관리 B
            "G": okosc_ws.Cells(r, 7).Value,    # 택배관리 C
            "H": okosc_ws.Cells(r, 8).Value,    # 택배관리 D
            "I": okosc_ws.Cells(r, 9).Value,    # 택배관리 E
        })
        r += 1

    if not rows_data:
        raise ValueError("통합문서에 데이터가 없습니다.")

    # ── 택배관리 파일에 붙여넣기 (openpyxl) ──────────────────────────────────
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # 기존 데이터 마지막 행 (A~E 중 하나라도 값 있는 행)
    last_row = utils.get_last_data_row(ws, check_cols=(1, 2, 3, 4, 5))
    new_start = last_row + 1

    for i, rd in enumerate(rows_data):
        target_row = new_start + i
        ws.cell(row=target_row, column=1).value = rd["E"]   # A
        ws.cell(row=target_row, column=2).value = rd["F"]   # B
        ws.cell(row=target_row, column=3).value = rd["G"]   # C
        ws.cell(row=target_row, column=4).value = rd["H"]   # D
        ws.cell(row=target_row, column=5).value = rd["I"]   # E
        ws.cell(row=target_row, column=6).value = config.DELIVERY_F_VALUE   # F
        ws.cell(row=target_row, column=7).value = config.DELIVERY_G_VALUE   # G
        ws.cell(row=target_row, column=8).value = config.DELIVERY_H_VALUE   # H
        ws.cell(row=target_row, column=11).value = rd["A"]  # K
        ws.cell(row=target_row, column=12).value = rd["C"]  # L
        ws.cell(row=target_row, column=13).value = rd["D"]  # M

    new_end = new_start + len(rows_data) - 1
    wb.save(xlsx_path)
    wb.close()
    return new_start, new_end


def step9_normalize_id_columns(xlsx_path: str, start_row: int, end_row: int):
    """
    D/E열 번호 정규화:
    - 둘 다 번호 있고 다름 → 그대로 유지
    - 하나만 번호 있음 → D열로 이동, E열 공란
    """
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    for row in range(start_row, end_row + 1):
        d_val = ws.cell(row=row, column=4).value
        e_val = ws.cell(row=row, column=5).value

        d_has = d_val not in (None, "")
        e_has = e_val not in (None, "")

        if d_has and e_has:
            # 둘 다 있으면 그대로
            pass
        elif e_has and not d_has:
            # E만 있으면 D로 이동
            ws.cell(row=row, column=4).value = e_val
            ws.cell(row=row, column=5).value = None
        # D만 있거나 둘 다 없으면 변경 없음

    wb.save(xlsx_path)
    wb.close()


def step10_11_paste_iksan_data(xlsx_path: str):
    """
    익산대장 L열 녹색 계열 셀 → 택배관리 마지막 행에 추가.
    이름(A), 전화(D), 주소(C), F=1, G=4400, H=한약
    각 셀 형식: "이름 전화번호 주소" (한 셀에 모두 포함)
    """
    iksan_path = utils.find_iksan_file()

    people = utils.get_iksan_green_cells(iksan_path)

    if not people:
        # 녹색 셀이 없으면 건너뜀
        return

    # 택배관리 파일에 추가
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    last_row = utils.get_last_data_row(ws, check_cols=(1, 3, 4))
    append_start = last_row + 1

    for i, (name, phone, addr) in enumerate(people):
        r = append_start + i
        ws.cell(row=r, column=1).value = name   # A: 이름
        ws.cell(row=r, column=3).value = addr   # C: 주소
        ws.cell(row=r, column=4).value = phone  # D: 전화번호
        ws.cell(row=r, column=6).value = config.DELIVERY_F_VALUE   # F
        ws.cell(row=r, column=7).value = config.DELIVERY_G_VALUE   # G
        ws.cell(row=r, column=8).value = config.DELIVERY_H_VALUE   # H

    wb.save(xlsx_path)
    wb.close()


def step13_16_upload_rosen(driver, xlsx_path: str):
    """
    로젠택배 사이트 → 예약관리-주문등록/출력(복수건) → 파일 업로드.
    스텝 16은 사람 검토이므로 함수 내에서 대기하지 않음.
    """
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC

    # 로젠택배 탭으로 이동하거나 새 탭 열기
    rosen_found = False
    for handle in driver.window_handles:
        driver.switch_to.window(handle)
        if "ilogen" in driver.current_url:
            rosen_found = True
            break
    if not rosen_found:
        driver.execute_script("window.open(arguments[0]);", config.ROSEN_URL)
        driver.switch_to.window(driver.window_handles[-1])

    wait = WebDriverWait(driver, config.PAGE_LOAD_TIMEOUT)

    # 예약관리 메뉴 클릭
    menu = wait.until(EC.element_to_be_clickable(
        (By.XPATH, "//*[contains(text(),'예약관리')]")
    ))
    menu.click()
    time.sleep(0.4)

    # 주문등록/출력(복수건) 클릭
    sub = wait.until(EC.element_to_be_clickable(
        (By.XPATH,
         "//*[contains(text(),'주문등록') and contains(text(),'복수건')] | "
         "//*[contains(text(),'복수건')]")
    ))
    sub.click()
    time.sleep(1)

    # 파일 업로드 input 찾기
    file_input = wait.until(EC.presence_of_element_located(
        (By.XPATH, "//input[@type='file']")
    ))
    file_input.send_keys(os.path.abspath(xlsx_path))
    time.sleep(0.5)


# ═══════════════════════════════════════════════════════════════════════════════
#  tkinter GUI 앱
# ═══════════════════════════════════════════════════════════════════════════════

class Auto1App:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("자동화 1번 - 택배 주문 처리")
        self.root.geometry("480x540")
        self.root.resizable(True, True)

        utils.set_root(root)

        self._q: queue.Queue = queue.Queue()
        self._running = False
        self._cancel = False
        self._xlsx_path = None

        self._build_ui()
        self._poll_queue()

    # ── UI 구성 ──────────────────────────────────────────────────────────────

    def _build_ui(self):
        top = tk.Frame(self.root, bg="#2C3E50", padx=10, pady=8)
        top.pack(fill="x")
        tk.Label(top, text="자동화 1번 - 택배 주문 처리",
                 font=("맑은 고딕", 13, "bold"),
                 bg="#2C3E50", fg="white").pack()

        self._status_var = tk.StringVar(value="■ 대기 중")
        tk.Label(self.root, textvariable=self._status_var,
                 font=("맑은 고딕", 10), fg="#2980B9",
                 anchor="w", padx=10).pack(fill="x", pady=(6, 0))

        log_frame = tk.Frame(self.root)
        log_frame.pack(fill="both", expand=True, padx=10, pady=4)
        self._log = scrolledtext.ScrolledText(
            log_frame, height=18, font=("Consolas", 9),
            state="disabled", bg="#FAFAFA"
        )
        self._log.pack(fill="both", expand=True)

        btn_frame = tk.Frame(self.root, pady=8)
        btn_frame.pack()
        self._run_btn = tk.Button(
            btn_frame, text="  시 작  ",
            command=self._start,
            font=("맑은 고딕", 11, "bold"),
            bg="#27AE60", fg="white", padx=18, pady=6,
            relief="flat", cursor="hand2"
        )
        self._run_btn.pack(side="left", padx=6)
        tk.Button(
            btn_frame, text="  닫 기  ",
            command=self.root.destroy,
            font=("맑은 고딕", 11),
            padx=18, pady=6,
            relief="flat", cursor="hand2"
        ).pack(side="left", padx=6)

    def _log_write(self, msg: str):
        self._log.config(state="normal")
        ts = datetime.now().strftime("%H:%M:%S")
        self._log.insert("end", f"[{ts}] {msg}\n")
        self._log.see("end")
        self._log.config(state="disabled")

    def _poll_queue(self):
        while not self._q.empty():
            kind, payload = self._q.get_nowait()
            if kind == "log":
                self._log_write(payload)
            elif kind == "status":
                self._status_var.set(payload)
            elif kind == "done":
                self._running = False
                self._run_btn.config(state="normal")
        self.root.after(80, self._poll_queue)

    def _put(self, kind: str, payload: str):
        self._q.put((kind, payload))

    def _log_msg(self, msg: str):
        self._put("log", msg)
        self._put("status", f"▶ {msg}")

    # ── 실행 ─────────────────────────────────────────────────────────────────

    def _start(self):
        if self._running:
            return
        self._running = True
        self._cancel = False
        self._run_btn.config(state="disabled")
        t = threading.Thread(target=self._run_all, daemon=True)
        t.start()

    def _run_all(self):
        import pythoncom
        pythoncom.CoInitialize()
        driver = None
        try:
            # 1. 다운로드
            self._log_msg("1단계: 홈페이지에서 택배 엑셀 다운로드 중...")
            driver = utils.connect_chrome()
            xls_path = step1_download_excel(driver)
            self._log_msg(f"  ✓ 다운로드: {os.path.basename(xls_path)}")

            # 2. XLS → XLSX
            self._log_msg("2단계: XLS → XLSX 변환 중...")
            xlsx_path = step2_convert_xls(xls_path)
            self._xlsx_path = xlsx_path
            self._log_msg(f"  ✓ 변환 완료: {os.path.basename(xlsx_path)}")

            # 3. 1행 삭제
            self._log_msg("3단계: 1행 삭제 중...")
            step3_remove_first_row(xlsx_path)
            self._log_msg("  ✓ 완료")

            # 4. G열 정리
            self._log_msg("4단계: G열 '즉납' → '4400' 교체 중...")
            step4_clean_column_g(xlsx_path)
            self._log_msg("  ✓ 완료")

            # 5. OKOSC 자동화
            self._log_msg("5단계: OKOSC에서 택배목록 가져오는 중...")
            okosc_wb = step5_automate_okosc()
            self._log_msg("  ✓ 통합문서 로드 완료")

            # 6-8. 데이터 붙여넣기
            self._log_msg("6~8단계: 통합문서 데이터 → 택배관리 붙여넣기 중...")
            new_start, new_end = step6_7_8_paste_okosc_data(xlsx_path, okosc_wb)
            self._log_msg(f"  ✓ {new_end - new_start + 1}행 추가됨 ({new_start}~{new_end}행)")

            # 9. D/E열 정규화
            self._log_msg("9단계: D/E열 번호 정규화 중...")
            step9_normalize_id_columns(xlsx_path, new_start, new_end)
            self._log_msg("  ✓ 완료")

            # 10-11. 익산대장
            self._log_msg("10~11단계: 익산대장 데이터 추가 중...")
            step10_11_paste_iksan_data(xlsx_path)
            self._log_msg("  ✓ 완료")

            # 12. 사람 검토
            self._log_msg("12단계: 사람 검토 대기 중...")
            ok = utils.human_review_dialog(
                "검토 요청",
                f"택배관리 파일을 확인해 주세요.\n\n"
                f"파일: {os.path.basename(xlsx_path)}\n\n"
                f"내용이 맞으면 '계속 진행'을 눌러주세요."
            )
            if not ok:
                self._log_msg("사용자가 중단했습니다.")
                return

            # 13-15. 로젠택배 업로드
            self._log_msg("13~15단계: 로젠택배 파일 업로드 중...")
            step13_16_upload_rosen(driver, xlsx_path)
            self._log_msg("  ✓ 파일 업로드 완료")

            # 16. 사람 검토 (오류 열 확인)
            self._log_msg("16단계: 오류 열 확인 요청...")
            utils.human_review_dialog(
                "오류 확인",
                "로젠택배 페이지의 '오류' 열에 오류가 없는지 확인해 주세요.\n\n"
                "확인이 완료되면 '계속 진행'을 눌러주세요.",
                ok_text="확인 완료"
            )

            self._log_msg("━━━ 자동화 1번 완료 ━━━")
            self._put("status", "✅ 완료")

        except Exception as e:
            self._log_msg(f"오류 발생: {e}")
            self._put("status", "❌ 오류 발생")
            self.root.after(0, lambda e=e: messagebox.showerror(
                "오류", str(e), parent=self.root
            ))
        finally:
            pythoncom.CoUninitialize()
            self._put("done", "")


# ═══════════════════════════════════════════════════════════════════════════════
#  진입점
# ═══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    root = tk.Tk()
    app = Auto1App(root)
    root.mainloop()
